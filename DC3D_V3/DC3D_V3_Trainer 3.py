# DeepClean Trainer v1.2.3
# Build created on Wednesday May 6th 2023
# Author: Adill Al-Ashgar
# University of Bristol
# adillwmaa@gmail.co.uk

"""
Possible improvements:

### REFACTOR: func: quantify_loss_performance ~ Appends the average performance metrics for the batch to the corresponding list of average metrics results for each epoch which is tracked eleshwehre in the enhanced performance tracking system and needs simplifying and condensing!!!

### REFACTOR: func: belief_telemetry ~ This function creates histogram plots of the pixel values recorded by the belief telemtry system [which needs renaming and reworking to simplify. (should be reconstruction thresholding telemtetry?)]  whihc records the values directly out of the netwrok before our reconstruction thresholding in the custom renomalisation is applied. This is important to keep ttrack of what is occusring before our iytput processing as it may be hiding errors.

### REFACTOR: func: create_settings_dict ~ Terrible, nneds compleate clean up and overhaul

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 

### REFACTOR: func:  ~ 




### ~~~~~ FIX BUG : xlsxwriter.exceptions.InvalidWorksheetName: Excel worksheet name 'T_fullframe_weighting_0.5 direct' must be <= 31 chars.

### ~~~~~ USe seeding value to controll data spareness and noise addition so that thay can be set to fixed

### ~~~~~ [DONE!] Create new unseen dataset for performance analysis on hyperparameter optimisation, could have a switch to make it use the val/test data?

### ~~~~~ [DONE!] Record the time taken per epoch, and then alongside the loss vs epoch plot (which dosent show time, so 1 epoch that takes 2 hrs is same as one that takes 1 min) plot loss vs time as a seperate plot

### ~~~~~ [DONE!] Add user controll to overide double precision processing

### ~~~~~ [DONE!] Improve the pixel telemtry per epoch by adding a dotted green li9ne indicating the true number of signal points\

### ~~~~~ Attach the new masking optimised normalisation check if it need a corresponding renorm

### ~~~~~ [DONE!] Make sure that autoecoder Encoder and Decoder are saved along with model in the models folder 

### ~~~~~ Add the new performance metrics per epoch tot he history da dictionary to clean up??

### ~~~~~ [DONE!] Fix memory leak in testing function loss calulation

### ~~~~~ [DONE!] Investigate and fix memory leak in plotting function

### ~~~~~ [DONE!] Reduce memory usage in loss calulation by removing lists

### ~~~~~ clean up the perforance loss plotting metircs calulation section, move to external script?

### ~~~~~ [DONE!] Ground plots and std devs in physical units

### ~~~~~ Clean up the performance loss plottsing code, it is too long and unwieldy, move it to an external file and load as a function

### ~~~~~~ [DONE!] Allow normalisation/renorm to be bypassed, to check how it affects results 

### ~~~~~~ [DONE!] Find out what is going on with recon threshold scaling issue

### ~~~~~~ [DONE!] fix noise adding to the data, it is not working as intended, need to retain clean images for label data 

### ~~~~~ Update the completed prints to terminal after saving things to only say that if the task is actually completeted, atm it will do it regardless, as in the error on .py save 

### ~~~~~ [DONE!] MUST SEPERATE 3D recon and flattening from the normalisation and renormalisation

### ~~~~~ #Noticed we need to update and cleanup all the terminal printing during the visulisations, clean up the weird line spaces and make sure to print what plot is currntly generating as some take a while and the progress bar doesn’t let user know what plot we are currently on

### ~~~~~ Update the model and data paths to folders inside the root dir so that they do not need be defined, and so that people can doanload the git repo and just press run without setting new paths etc 

### ~~~~~ fix epoch numbering printouts? they seem to report 1 epoch greater than they should

### ~~~~~ clear up visulisations

### ~~~~~ Functionalise things

### ~~~~~ Clean up custom autoencoder.py file saver terminal prints left over from debugging

### ~~~~~ [DONE!] Fix Val loss save bug

### ~~~~~ [DONE!] Custom MSE loss fucntion with weighting on zero vals to solve class imbalence

### ~~~~~ [DONE!] Move things to other files (AE, Helper funcs, Visulisations etc)

### ~~~~~ [DONE!] Fix reconstruction threshold, use recon threshold to set bottom limit in custom normalisation

### ~~~~~ [DONE!] Turn plot or save into a function 

### ~~~~~ [DONE!] Add in a way to save the model after each epoch, so that if the program crashes we can still use the last saved model

### ~~~~~ [DONE!] Find way to allow user to exit which is non blocking 

### ~~~~~ [DONE!] Train on labeld data which has the fill line paths in labels and then just points on line in the raw data?

### ~~~~~ Add arg to plot or save function that passes a test string to print for the plot generating user notice ratehr than the generic one used each time atm 

### ~~~~~ [DONE!] change telemetry variable name to output_pixel_telemetry

### ~~~~~ [DONE!] Fix this " if plot_higher_dim: AE_visulisation(en...)" break out all seperate plotting functions
    
### ~~~~~ adapt new version for masking - DeepMask3D 

### ~~~~~ sort out val, test and train properly

### ~~~~~ FC2_INPUT_DIM IS NOT USED!! This would be extremely useful. ?? is this for dynamic layer sizing?

### ~~~~~ Update all layer activation tracking from lists and numpy to torch tensors throughout pipeline for speed

### ~~~~~ Add in automatic Enc/Dec layer size calulations

### ~~~~~ [DONE!] Search for and fix errors in custom norm an renorm

### ~~~~~ [DONE!] Seperate and moularise renorm and 3D reconstruction

### ~~~~~ Create flatten module in main body so noise can be added to the 3D cube rather than slicewise

### ~~~~~ [DONE!] Add way to compress the NPZ output as filesize is to large ! ~3Gb+

### ~~~~~ [DONE!] Add all advanced program settings to end of net summary txt file i.e what typ eof normalisation used etc, also add th enam eof the autoencoder file i.e AE_V1 etc from the module name 

### ~~~~~ [DONE!] update custom mse loss fucntion so that user arguments are set in settings page rather than at function def by defualts i.e (zero_weighting=1, nonzero_weighting=5)

### ~~~~~ [DONE!] could investigate programatically setting the non_zero weighting based on the ratio of zero points to non zero points in the data set which would balance out the two classes in the loss functions eyes

### ~~~~~ [DONE!] Add way for program to save the raw data for 3d plots so that they can be replotted after training and reviewed in rotatable 3d 

### ~~~~~ [DONE!] Check if running model in dp (fp64) is causing large slow down???

### ~~~~~ Allow seperate loss fucntion for testing/validation phase?

### ~~~~~ Properly track the choices for split loss funcs in txt output file 

### ~~~~~ Explicitlly pass in split_loss_functions to the split custom weigted func atm is not done to simplify the code but is not ideal

### ~~~~~ [DONE!] Update noise points to take a range as input and randomly select number for each image from the range

### ~~~~~ [DONE!] Add fcuntion next to noise adder that drops out pixels, then can have the labeld image with high signal points and then dropout the points in the input image to network so as to train it to find dense line from sparse points!

### ~~~~~ [DONE!] Add plots of each individual degradation step rathert than just all shown on one (this could be done instead of the current end of epoch 10 plots or alongside)

### ~~~~~ colour true signal points red in the input distroted image so that viewer can see the true signal points and the noise added

### ~~~~~ [DONE!] add masking directly to the trainer so we can see masked output too 

### ~~~~~  Label loss plots y axis programatically based on user loss function selection
"""
 
# NOTE to users: First epoch will always run slower when using a new dataset or after a computer restart as system memory is being trained, subsequent epochs should take ~50% of the time of the first epoch
# NOTE to users: The 'nonzero_weighting' parameter is a great way to adjust the sensetivity of the training result. Values around 0.1 will be very cautious in predicting hits, whilst moving to around 1.0 will be very confident in predicting hits. This is a great way to adjust the sensetivity of the model to your needs. Low values are better for direct net output, whilst higher values are better for masking output.





#%% - User Inputs
dataset_title = "RDT 10K MOVE"# "RF_5K"#"PDT 10K" #"RDT 10K MOVE" #'RDT 500K 1000ToF' #"RDT 10K MOVE" #"RDT 50KM"# "Dataset 37_X15K Perfect track recovery" #"Dataset 24_X10Ks"           #"Dataset 12_X10K" ###### TRAIN DATASET : NEED TO ADD TEST DATASET?????
model_save_name = "DEEPER AUTOENC TEST"#'Parabola6 no norm to loss' #"T2"#"RDT 500K 1000ToF timed"#"2023 Testing - RDT100K n100"#"2023 Testing - RDT10K NEW" #"RDT 100K 30s 200n Fixed"#"RDT 50KM tdim1000 AE2PROTECT 30 sig 200NP LD10"     #"D27 100K ld8"#"Dataset 18_X_rotshiftlarge"

TORCHSIM_data = False #!!!!!!!!!!!!!!!!!!!!!!

xdim = 88   # Currently useless
ydim = 128  # Currently useless
time_dimension = 1000                        # User controll to set the number of time steps in the data
channels = 1                                # User controll to set the number of channels in the data

#%% - Training Hyperparameter Settings
num_epochs = 2000                            # User controll to set number of epochs (Hyperparameter)
batch_size = 10 #6 looks good                              # User controll to set batch size - number of Images to pull per batch (Hyperparameter) 
learning_rate = 0.0001                           # User controll to set optimiser learning rate (Hyperparameter)
optim_w_decay = 1e-05                        # User controll to set optimiser weight decay for regularisation (Hyperparameter)
BOOST = 1

latent_dim = 50     #NOTE: Tested!                         # User controll to set number of nodes in the latent space, the bottleneck layer (Hyperparameter)
fc_input_dim = 512                         # User controll to set number of nodes in the fc2 layer (Hyperparameter)
dropout_prob = 0.2                           # [NOTE Not connected yet] User controll to set dropout probability (Hyperparameter)
reconstruction_threshold = 0.5               # MUST BE BETWEEN 0-1  #Threshold for 3d reconstruction, values below this confidence level are discounted

#%% - Dataset Settings
train_test_split_ratio = 0.9                 # User controll to set the ratio of the dataset to be used for training (Hyperparameter)
val_set_on = False                           # User controll to set if a validation set is used
val_test_split_ratio = 0.9                   # This needs to be better explained its actually test_val ration ratehr than oterh way round     # [NOTE LEAVE AT 0.5, is for future update, not working currently] User controll to set the ratio of the non-training data to be used for validation as opposed to testing (Hyperparameter)

#%% - Loss Function Settings
loss_vs_sparse_img = False    #NOTE: Tested!                # User controll to set if the loss is calculated against the sparse image or the full image (Hyperparameter)
loss_function_selection = 0                  # Select loss function (Hyperparameter): 0 = ACBMSE, 1 = ffACBMSE, 2 = ACBMSE3D, 3 = torch.nn.MSELoss(), 4 = torch.nn.BCELoss(), 5 = torch.nn.L1Loss(), 6 = ada_SSE_loss, 7 = ada_weighted_custom_split_loss, 8 = weighted_perfect_reconstruction_loss

# Below weights only used if loss func set to 0, 1 or 6 aka ACBMSE or split loss varients
zero_weighting = 1                        # User controll to set zero weighting for ACBMSE (Hyperparameter)
nonzero_weighting = 1#0.4 #0.4                        # User controll to set non zero weighting for ACBMSE (Hyperparameter)
# Only used for ffACBMSE along with above two settings
fullframe_weighting = 1.5 #1.5                      # User controll to set full frame weighting for ffACBMSE (Hyperparameter)
ff_loss = 'mse'

# Below only used if loss func set to 6 aka ada_weighted_custom_split_loss
zeros_loss_choice = 1                        # Select loss function for zero values (Hyperparameter): 0 = Maxs_Loss_Func, 1 = torch.nn.MSELoss(), 2 = torch.nn.BCELoss(), 3 = torch.nn.L1Loss(), 4 = ada_SSE_loss
nonzero_loss_choice = 1                      # Select loss function for non zero values (Hyperparameter): 0 = Maxs_Loss_Func, 1 = torch.nn.MSELoss(), 2 = torch.nn.BCELoss(), 3 = torch.nn.L1Loss(), 4 = ada_SSE_loss

#%% - Image Preprocessing Settings  (when using perfect track images as labels)
signal_points = 30                          # User controll to set the number of signal points to add
noise_points =  200#0#100                         # User controll to set the number of noise points to add

x_std_dev = 0  #mm NOTE ADAPTed TO PHYSICAL                             # User controll to set the standard deviation of the detectors error in the x axis
y_std_dev = 0  #mm NOTE ADAPTed TO PHYSICAL                             # User controll to set the standard deviation of the detectors error in the y axis
tof_std_dev = 0  #ns NOTE ADAPTed TO PHYSICAL                           # User controll to set the standard deviation of the detectors error in the time of flight 

#%% - Pretraining settings
start_from_pretrained_model = False         # If set to true then the model will load the pretrained model and optimiser state dicts from the path below
load_pretrained_optimser = True             # Only availible if above is set to true - (pretrain seems to perform better if this is set to true)
pretrained_model_path = 'N:/Yr 3 Project Results/RICHFAST TEST 5 - Training Results/Model_Deployment/RICHFAST TEST 5 - Model + Optimiser State Dicts.pth'      # Specify the path to the saved full state dictionary for pretraining

#%% - Normalisation Settings 
masking_optimised_binary_norm = False       # If set to true then the model will use the binary normalisation method optimised for masking output. Otherwise will use the gaped custom normalisation optimised for the direct network output

#%% - Plotting Control Settings
print_every_other = 1                       # [default = 2] 1 is to save/print all training plots every epoch, 2 is every other epoch, 3 is every 3rd epoch etc
plot_or_save = 1                            # [default = 1] 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)
num_to_plot = 10
save_all_raw_plot_data = True               # [default = False] If set to true then all raw data for plots is saved to disk for replotting and analysis later

#%% - New beta feature settings 
double_precision = False #NOTE: Tested!     # [default = False] If set to true then the model will use double precision floating point numbers for all calculations, otherwise will use single precision
shuffle_train_data = True #NOTE: Tested!    # If set to true then the model will shuffle the training data each epoch, otherwise will not shuffle

# TRUN THESE TWO FLAGS INTO ONE SINGLE VARIABLE WHERE None WILL HAVE NO TIMEOUT AND A VALUE WILL SET THE TIMEOUT
timeout_training = False                    # If set to true then the model will stop training after the timeout time has been reached
timeout_time = 720                          # Time in minuits to wait before stopping training if timeout_training is set to true`

record_weights = False                      # If set to true then the model will record the weights of the network at the end of each epoch
record_biases = False                       # If set to true then the model will record the biases of the network at the end of each epoch
record_activity = False                     # If set to true then the model will record the activations of the network at the end of each epoch
compress_activations_npz_output = False     # Compresses the activity file above for smaller file size but does increase loading and saving times for the file. (use if low on hdd space)

### NEW PHYSICAL GROUNDING FOR UNITS
use_physical_values_for_plot_axis = False    # If false then axis are label by pixel indicies, if true then axis are labelled by physical units
x_length = 800 #mm
y_length = 1500 #mm
time_length = 1000 #ns
time_scale = time_dimension / time_length # ns per t pixel
x_scale = xdim / x_length # mm per x pixel
y_scale = ydim / y_length # mm per y pixel
physical_scale_parameters = [x_scale, y_scale, time_scale]

#%% - Advanced Visulisation Settings
plot_train_loss = True                      # [default = True]       
plot_test_loss = True                       # [default = True]
plot_validation_loss = True                 # [default = True]               
plot_time_loss = True                       # [default = True]
plot_detailed_performance_loss = True       # Plots ssim nmi etc for each epoch 
use_tensorboard = False
plot_live_time_loss = True                  # [default = True] Generate plot of live training loss vs time during trainig which is overwritten each epoch, this is useful for seeing how the training is progressing
plot_live_training_loss = True              # [default = True] Generate plot of live training loss vs epoch during trainig which is overwritten each epoch, this is useful for seeing how the training is progressing
comparative_live_loss = True                # [default = True] Adds comparative lines to the live plots, the models for comparison are selected below
slide_live_plot_size = 0                    # [default = 0] Number of epochs to show on the live plot (if set to 0 then will show all epochs)
plot_comparative_loss = True

comparative_loss_titles = ["S1 10K", "S1 100K", "S1 500K", "S2 10K"]
comparative_loss_paths = [r'N:\Yr 3 Project Results\RDT 10K 1000ToF timed - Training Results\\',   # Settings V1
                          r'N:\Yr 3 Project Results\RDT 100K 1000ToF timed - Training Results\\',  # Settings V1
                          r'N:\Yr 3 Project Results\RDT 500K 1000ToF timed - Training Results\\',  # Settings V1
                          r'N:\Yr 3 Project Results\RDT 10K S2 - Training Results\\'  # Settings V2
                          ] 
                          
plot_pixel_threshold_telemetry = True     # [default = False] # Update name to pixel_cuttoff_telemetry    #Very slow, reduces net performance by XXXXXX%
plot_pixel_difference = False #BROKEN     # [default = True]          
plot_latent_generations = True            # [default = True]              
plot_higher_dim = False                   # [default = True]  
plot_normalised_radar = False
plot_Graphwiz = True                      # [default = True]       

#%% - Advanced Debugging Settings
print_encoder_debug = False                     # [default = False]  
print_decoder_debug = False                     # [default = False] 
print_network_summary = False                   # [Default = False] Prints the network summary to terminal
print_partial_training_losses = False           # [Default = True] Prints the training loss for each batch in the epoch

debug_noise_function = False                    # [default = False]  
debug_loader_batch = False                      # SAFELY REMOVE THIS PARAM!!!  #(Default = False) //INPUT 0 or 1//   #Setting debug loader batch will print to user the images taken in by the dataoader in this current batch and print the corresponding labels
debug_model_exporter  = False                   # [default = False]

full_dataset_integrity_check = False            # [Default = False] V slow  #Checks the integrity of the dataset by checking shape of each item as opposed to when set to false which only checks one single random file in the dataset
full_dataset_distribution_check = False         # [Default = False] V slow  #Checks the distribution of the dataset , false maesn no distributionn check is done

seeding_value = 10 #None                            # [Default = None] None gives no seeeding to RNG, if the value is set this is used for the RNG seeding for numpy, and torch libraries

#%% Hyperparameter Optimisation Settings  #######IMPLEMENT!!!
optimise_hyperparameter = False              # User controll to set if hyperparameter optimisation is used
hyperparam_to_optimise = 'zero_weighting'      # User controll to set which hyperparameter to optimise  - options are: 'batch_size', 'learning_rate', 'optim_w_decay', 'dropout_prob', 'loss_function_selection', 'conv_layers', 'conv_filter_multiplier', 'latent_dim'
set_optimisiation_list_manually = [0.3, 0.5, 0.7, 1.0, 1.3]   #  set this param = to your list i.e [[12, 120], 35, 87]

# Simple Performance Measure
print_validation_results = True            # User controll to set if the validation results are printed to terminal 
plot_training_time = True                   # User controll to set if the training time is plotted 

# Full Performance Analysis - Performed in addition to and seperatly from the validation stage for automatic data analysis
perf_analysis_num_files = 5000   # number of files to test
perf_analysis_plot = 100       # The number of results to print imshow plots for for each model tested, set to False for none
perf_analysis_dataset_dir = (r"N:\\Yr 3 Project Datasets\\PERF VALIDATION SETS\\40K 100N 30S\\")   # directory of dataset to test - (Best to use totally unseen data files that are not contianed within the train, test or validation sets)
debug_hpo_perf_analysis = False

#%% - HACKS NEED FIXING!!!
if print_every_other > num_epochs:   #protection from audio data not being there to save if plot not genrated - Can fix by moving audio genration out of the plotting function entirely and only perform it once at the end wher eit is actually saved.
    print_every_other = num_epochs

history_da = {'train_loss':[], 'test_loss':[], 'val_loss':[], 'HTO_val':[], 'training_time':[]} # needs better placement???
max_epoch_reached = 0 # in case user exits before end of first epoch 

# Create a dictionary to store the activations
activations = {}
weights_data = {}
biases_data = {}

#%% - Data Path Settings
data_path = "N:\Yr 3 Project Datasets\\"
results_output_path = "N:\Yr 3 Project Results\\"
results_output_path_1 = results_output_path  # HACK FOR HYPERPARAM OPTIMISATION FIX IT!!!


#%% - Create full settings dictionary !!! NEEDS CLEANUP

"""
if optimise_hyperparameter:
    settings_path = results_output_path_1 + model_save_name + "_" + hyperparam_to_optimise + f" Optimisation\\"
else:
    settings_path = results_output_path + model_save_name + " - Training Results/"
#create settings path if not exists
os.makedirs(settings_path, exist_ok=True)
full_settings_dictionary = create_settings_dict(settings_path + "full_settings_dictionary")

"""




#%% - Load in Dependencies
# External Libraries
import os
import time     # Used to time the training loop
import json
import torch
import pickle
import datetime 
import torchvision 
import numpy as np  
import pandas as pd
from tqdm import tqdm  # Progress bar
import matplotlib.cm as cm
from functools import partial
from torchinfo import summary # function to get the summary of the model layers structure, trainable parameters and memory usage
import matplotlib.pyplot as plt     
from torchvision import transforms  
from matplotlib.ticker import FuncFormatter
from torch.utils.tensorboard import SummaryWriter, FileWriter 


# - Helper functions
from Helper_files.Helper_Functions import *
from Helper_files.Robust_model_exporter_V1 import Robust_model_export                   # Function to export the raw .py file that contains the autoencoder class
from Helper_files.System_Information_check import get_system_information                # Function to get the host system performance specs of the training machine
from Helper_files.Dataset_Integrity_Check_V1 import dataset_integrity_check             # Function to check the integrity of the datasets values
from Helper_files.Dataset_distribution_tester_V1 import dataset_distribution_tester     # Function to check the distribution of the datasets values
from Helper_files.AE_Visulisations import Generative_Latent_information_Visulisation, Reduced_Dimension_Data_Representations, Graphviz_visulisation, AE_visual_difference # Functions to visulise the autoencoders training progression
from Helper_files.ExcelExtractor import extract_data_to_excel # This is a custom function to extract data from excel files
from Helper_files.model_perf_analysis2 import run_full_perf_tests
from Helper_files.Image_Metrics import MSE, MAE, SNR, PSNR, NMSRE, SSIM, NomalisedMutualInformation, correlation_coeff, compare_images_pixels
#from Helper_files.Export_User_Settings import create_settings_dict

# - Loss Functions
from Loss_Functions.Loss_Fn_Classes import *

# - Autoencoder
from Autoencoders.DC3D_Autoencoder_V1_Protected2_3 import Encoder, Decoder   





#%% - Load in Comparative Live Loss Data

if comparative_live_loss:
    #history_da, epoch_times = load_comparative_data(comparative_loss_paths, plot_live_training_loss, plot_live_time_loss)

    
    comparative_history_da = []
    comparative_epoch_times = []
    for loss_path in comparative_loss_paths:
        #load pkl file into dictionary
        if plot_live_training_loss or plot_live_time_loss:
            with open(loss_path + '\\Raw_Data_Output\\history_da_dict.pkl', 'rb') as f:
                comparative_history_da.append(pickle.load(f))
        if plot_live_time_loss:
            with open(loss_path + '\\Raw_Data_Output\\epoch_times_list_list.csv', 'rb') as f:
                # load the data from the csv file called f into a list 
                comparative_epoch_times.append(np.loadtxt(f, delimiter=',').tolist())
               


          


#%% DC3D Special Functions
# Special normalisation for pure masking
def mask_optimised_normalisation(data):
    """
    Normalisation function for pure masking output, trhe function takes any non zero valu eto 1 and any zero value is left as is. [EXPLAIN WHY!!!!!]

    Args:
        data (torch tensor): The input data to be normalised.

    Returns:
        data (torch tensor): The normalised data.
    
    """
    data = torch.where(data > 0, 1.0, 0.0)
    return data

# Custom normalisation function
def gaped_normalisation(data, reconstruction_threshold, time_dimension=100):
    """
    Normalisation function that normalised values in range [0 < value <= 'time_dimension'] to new range of [reconstruction_threshold' < norm_value <= 1], with a gap at the bottom of the range to allow for the reconstruction threshold to be applied later. All zero values are left untouched. This is used for the direct network output.
    
    Args:
        data (torch tensor): The input data to be normalised. [WARNING: the data must be (integer values? and) in the range [0 < value <= 'time_dimension'] or the normalisation will not work correctly]
        reconstruction_threshold (float): The threshold used in the custom normalisation, used to set the lower limit of the normalised values.
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the normalised values. Default = 100  

    Returns:
        data (torch tensor): The normalised data.
    """
    data = torch.where(data > 0, (((data / time_dimension) / (1 / (1 - reconstruction_threshold))) + reconstruction_threshold), 0 )
    return data

# Custom renormalisation function
def gaped_renormalisation_torch(data, reconstruction_threshold, time_dimension=100):
    """
    torch version of our Renormalisation function that renormalises values in range [reconstruction_threshold < value <= 1] to new range of [0 < renorm_value <= 'time_dimension'] removing any values that fell below the 'reconstruction_threshold' by setting thwm to zero. All zero values are left untouched. This is used for the direct network output.

    Args:
        data (torch tensor): The input data to be renormalised. [WARNING: the data must be (float values? and) in the range [reconstruction_threshold < value <= 1] or the renormalisation will not work correctly]
        reconstruction_threshold (float): The threshold used in the custom normalisation, used to set the lower limit of the normalised values.
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the normalised values. Default = 100

    Returns:
        data (torch tensor): The renormalised data.
    """

    data = torch.where(data > reconstruction_threshold, ((data - reconstruction_threshold)*(1 / (1 - reconstruction_threshold))) * (time_dimension), 0)
    return data

def gaped_renormalisation(data, reconstruction_threshold, time_dimension=100):
    """
    Numpy version of our Renormalisation function that renormalises values in range [reconstruction_threshold < value <= 1] to new range of [0 < renorm_value <= 'time_dimension'] removing any values that fell below the 'reconstruction_threshold' by setting thwm to zero. All zero values are left untouched. This is used for the direct network output.

    Args:
        data (np array): The input data to be renormalised. [WARNING: the data must be (float values? and) in the range [reconstruction_threshold < value <= 1] or the renormalisation will not work correctly]
        reconstruction_threshold (float): The threshold used in the custom normalisation, used to set the lower limit of the normalised values.
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the normalised values. Default = 100

    Returns:
        data (np array): The renormalised data.
    """
    data = np.where(data > reconstruction_threshold, ((data - reconstruction_threshold)*(1 / (1 - reconstruction_threshold))) * (time_dimension), 0)
    return data

# 3D Reconstruction function
def reconstruct_3D(*args):
    """
    3D reconstruction function that takes any number of 2D arrays created through our 3D to 2D with embedded ToF method and reconstructs them into a 3D arrays. 
    
    Args:
        *args (np array): Any number of 2D arrays to be reconstructed.

    Returns:
        results (np array): The reconstructed 3D arrays. Same number of arrays as input.
    """
    
    results = []
    for data in args:
        data_output = []
        for cdx, row in enumerate(data):
            for idx, num in enumerate(row):
                if num > 0:  
                    data_output.append([cdx, idx, num])
        results.append(np.array(data_output))

    return results

# Masking technique
def masking_recovery(input_image, recovered_image, time_dimension, print_result=False):
    """
    Applies my masking technique to the recovered image, this is essential and and condational for both the input and the recovered image as to weatehr a pixel is allowed to stay in the final output or not, which relies on the differing distortion profiles of the two images.

    Args:
        input_image (np array): The input image to be masked.
        recovered_image (np array): The recovered image to be masked.
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the normalised values.
        print_result (bool, optional): If set to true then the function will print a text string reporting the masking usefullness evaluation metrics. Defaults to False.

    Returns:
        result (np array): The masked recovered image.
    
    """
    raw_input_image = input_image.copy()
    net_recovered_image = recovered_image.copy()
    #Evaluate usefullness 
    # count the number of non-zero values
    masking_pixels = np.count_nonzero(net_recovered_image)
    image_shape = net_recovered_image.shape
    total_pixels = image_shape[0] * image_shape[1] * time_dimension
    # print the count
    if print_result:
        print(f"Total number of pixels in the timescan: {format(total_pixels, ',')}\nNumber of pixels returned by the masking: {format(masking_pixels, ',')}\nNumber of pixels removed from reconstruction by masking: {format(total_pixels - masking_pixels, ',')}")

    # use np.where and boolean indexing to update values in a
    mask_indexs = np.where(net_recovered_image != 0)
    net_recovered_image[mask_indexs] = raw_input_image[mask_indexs]
    result = net_recovered_image
    if result.shape != raw_input_image.shape:
        print("ERROR: Masking has failed, the recovered image is not the same shape as the input image")
    return result

#%% - Data Degradation Functions

# Function to add n noise points to each image in a tensor batch(must be used AFTER custom norm 
def add_noise_points_to_batch(input_image_batch, noise_points=100, reconstruction_threshold=0.5):
    """
    This function will take a batch of images and add noise_points number of random noise points to each image in the batch. Intended use case is adding noise points to the image after the custom normalisation is applied, the noise points added will be floats between recon_threshold and 1.

    Args:
        input_image_batch (torch tensor): The input image batch to be degraded. Shape [B, C, H, W]  
        noise_points (int): The number of noise points to add to each image in the batch
        reconstruction_threshold (float): The threshold used in the custom normalisation, used to set the lower limit of the noise point values. Default = 0.5
    
    Returns:
        image_batch (torch tensor): The degraded image batch. Shape [B, C, H, W]
    """

    image_batch = input_image_batch.clone()
    if noise_points > 0:
        #Find dimensions of input image 
        x_dim = image_batch.shape[2]
        y_dim = image_batch.shape[3]

        #For each image in the batch
        for image in image_batch:

            # Create a list of unique random x and y coordinates
            num_pixels = x_dim * y_dim
            all_coords = np.arange(num_pixels)
            selected_coords = np.random.choice(all_coords, noise_points, replace=False)
            x_coords, y_coords = np.unravel_index(selected_coords, (x_dim, y_dim))
            
            # Iterate through noise_points number of random pixels to noise
            for i in range(noise_points):

                # Add a random number between recon_threshold and 1 to the pixel 
                image[0][x_coords[i], y_coords[i]] = np.random.uniform(reconstruction_threshold, 1)

    return image_batch

# Function to add n noise points to each image in a tensor batch 
def add_noise_points_to_batch_prenorm(input_image_batch, noise_points=100, time_dimension=100):
    """
    This function will take a batch of images and add noise_points number of random noise points to each image in the batch. Intended use case is adding noise points to the image before the custom normalisation is applied, the noise points added will be integers between 1 and 'time_dimension' which is the max timestep.

    Args:
        input_image_batch (torch tensor): The input image batch to be degraded. Shape [B, C, H, W]  
        noise_points (int): The number of noise points to add to each image in the batch
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values. Default = 100

    Returns:
        image_batch (torch tensor): The degraded image batch. Shape [B, C, H, W]
    """

    image_batch = input_image_batch.clone()
    if noise_points > 0:
        #Find dimensions of input image 
        x_dim = image_batch.shape[2]
        y_dim = image_batch.shape[3]

        #For each image in the batch
        for image in image_batch:

            # Create a list of unique random x and y coordinates
            num_pixels = x_dim * y_dim
            all_coords = np.arange(num_pixels)
            selected_coords = np.random.choice(all_coords, noise_points, replace=False)  ##### NOTE: Change to using a torch random function instead of a numpy one
            x_coords, y_coords = np.unravel_index(selected_coords, (x_dim, y_dim))
            
            # Iterate through noise_points number of random pixels to noise
            for i in range(noise_points):

                # Add a random number between recon_threshold and 1 to the pixel 
                image[0][x_coords[i], y_coords[i]] = np.random.uniform(0, time_dimension)   ##### NOTE: Change to using a torch random function instead of a numpy one

    return image_batch

# Function to create sparse signal from a fully dense signal
def create_sparse_signal(input_image_batch, signal_points=2, linear=False):

    """
    This function will take a batch of images and randomly select signal_points number of non-zero values to keep in each image, zeroing out all other values i.e if signal_points = 2 then only 2 non-zero values will be kept in each image and all other non zero values will be zeroed out, effectivly simulating a very sparse signal
    
    Args:
        input_image_batch (torch tensor): The input image batch to be degraded. Shape [B, C, H, W]
        signal_points (int): The number of non-zero values to keep in each image
        linear (bool): If set to true then the signal points are linearly spaced across the signal, otherwise they are randomly selected. Default = False
    """

    # Take as input a torch tensor in form [batch_size, 1, x_dim, y_dim]]
    # Create a copy of the input image batch
    image_batch = input_image_batch.clone()

    # Flatten the image tensor
    flat_batch = image_batch.view(image_batch.size(0), -1)

    # Count the number of non-zero values in each image
    nz_counts = torch.sum(flat_batch != 0, dim=1)

    # Find the indices of the images that have more non-zero values than signal_points
    sparse_indices = torch.where(nz_counts > signal_points)[0]

    # For each sparse image, randomly select signal_points non-zero values to keep
    for idx in sparse_indices:
        # Find the indices of the non-zero values in the flattened image
        nz_indices = torch.nonzero(flat_batch[idx]).squeeze()

        # Randomly select signal_points non-zero values to keep
        if linear:
            kept_indices = torch.linspace(0, nz_indices.numel() - 1, steps=signal_points).long()
        else:
            kept_indices = torch.randperm(nz_indices.numel())[:signal_points]

        # Zero out all non-selected values
        nonkept_indices = nz_indices[~torch.isin(nz_indices, nz_indices[kept_indices])]
        flat_batch[idx, nonkept_indices] = 0

    # Reshape the flat tensor back into the original shape
    output_image_batch = flat_batch.view_as(image_batch)

    return output_image_batch

# Function to add shift in x, y and ToF to a true signal point due to detector resoloution
def simulate_detector_resolution(input_image_batch, x_std_dev, y_std_dev, tof_std_dev, x_scale, y_scale, time_scale, plot=False):
    """
    This function will add a random shift taken from a gaussain std deviation in x, y or ToF to each non-zero pixel in the image tensor to simulate detector resoloution limits

    Args:
        input_image_batch (torch tensor): The input image batch to be degraded. Shape [B, C, H, W]
        x_std_dev (float): The standard deviation of the gaussian distribution to draw the x shift from
        y_std_dev (float): The standard deviation of the gaussian distribution to draw the y shift from
        tof_std_dev (float): The standard deviation of the gaussian distribution to draw the ToF shift from
        x_scale (float): The physical scale of the x axis in mm per pixel to convert pixel values to physical values
        y_scale (float): The physical scale of the y axis in mm per pixel to convert pixel values to physical values
        time_scale (float): The physical scale of the ToF axis in ns per pixel to convert pixel values to physical values

    Returns:
        image_batch_all (torch tensor): The degraded image batch. Shape [B, C, H, W]
    """

    # Convert physical values to pixel values
    x_std_dev_pixels = x_std_dev / x_scale
    y_std_dev_pixels = y_std_dev / y_scale
    tof_std_dev_pixels = tof_std_dev / time_scale

    # Take as input a torch tensor in form [batch_size, 1, x_dim, y_dim]]
    # Create a copy of the input image batch
    image_batch_all = input_image_batch.clone()

    for idx, image_batch_andc in enumerate(image_batch_all):
        image_batch = image_batch_andc.squeeze()
        # Assume that the S2 image is stored in a variable called "image_tensor"
        x, y = image_batch.size()

        # For all the values in the tensor that are non zero (all signal points) adda random value drawn from a gaussian distribution with mean of the original value and std dev of ToF_std_dev so simulate ToF resoloution limiting
        image_batch[image_batch != 0] = image_batch[image_batch != 0] + torch.normal(mean=0, std=tof_std_dev_pixels, size=image_batch[image_batch != 0].shape)

        # Generate random values for shifting the x and y indices
        x_shift = torch.normal(mean=0, std=x_std_dev_pixels, size=(x, y), dtype=torch.float32)
        y_shift = torch.normal(mean=0, std=y_std_dev_pixels, size=(x, y), dtype=torch.float32)

        # Create a mask for selecting non-zero values in the image tensor
        mask = image_batch != 0

        # Apply the x and y shifts to the non-zero pixel locations
        new_x_indices = torch.clamp(torch.round(torch.arange(x).unsqueeze(1) + x_shift), 0, x - 1).long()
        new_y_indices = torch.clamp(torch.round(torch.arange(y).unsqueeze(0) + y_shift), 0, y - 1).long()
        shifted_image_tensor = torch.zeros_like(image_batch)
        shifted_image_tensor[new_x_indices[mask], new_y_indices[mask]] = image_batch[mask]

        if plot:
            plt.imshow(shifted_image_tensor, cmap='gray', vmin=0, vmax=100)
            plt.title('S')
            plt.show()

        image_batch_all[idx,0] = shifted_image_tensor
        
    return image_batch_all

def signal_degredation(signal_settings, image_batch, physical_scale_parameters):
    """
    Sequentially applies the differnt signal degredation functions to the input image batch and returns the output of each stage

    Args:
        signal_settings (list): A list of the signal degredation settings to be applied to the input image batch. Contains [signal_points_r, x_std_dev_r, y_std_dev_r, tof_std_dev_r, noise_points_r]
        image_batch (torch tensor): The input image batch to be degraded. Shape [B, C, H, W]
        physical_scale_parameters (list): A list of the physical scale parameters to be used to convert pixel values to physical values. Contains [x_scale, y_scale, time_scale]
            x_scale (float): The physical scale of the x axis in mm per pixel to convert pixel values to physical distance values
            y_scale (float): The physical scale of the y axis in mm per pixel to convert pixel values to physical distance values
            time_scale (float): The physical scale of the ToF axis in ns per pixel to convert pixel values to physical time values
    """
    x_scale, y_scale, time_scale = physical_scale_parameters
    signal_points_r, x_std_dev_r, y_std_dev_r, tof_std_dev_r, noise_points_r = signal_settings
    sparse_output_batch = create_sparse_signal(image_batch, signal_points_r)
    sparse_and_resolution_limited_batch = simulate_detector_resolution(sparse_output_batch, x_std_dev_r, y_std_dev_r, tof_std_dev_r, x_scale, y_scale, time_scale)
    noised_sparse_reslimited_batch = add_noise_points_to_batch_prenorm(sparse_and_resolution_limited_batch, noise_points_r, time_dimension)
    return sparse_output_batch, sparse_and_resolution_limited_batch, noised_sparse_reslimited_batch   

#%% - Network Hook Functions
def activation_hook_fn(module, input, output, layer_index):
    """
    This function will be called whenever a layer is called during the forward pass.
    It records the activations and saves them in the specified dictionary.
    """
    activations[layer_index] = output.detach()

def weights_hook_fn(module, input, output, layer_index):
    """
    This function will be called whenever a layer is called during the forward pass.
    It records the activations and saves them in the specified dictionary.
    """
    weights_data[layer_index] = module.weight.data.clone().detach()

def biases_hook_fn(module, input, output, layer_index):
    """
    This function will be called whenever a layer is called during the forward pass.
    It records the activations and saves them in the specified dictionary.
    """
    biases_data[layer_index] = module.bias.data.clone().detach()

def write_hook_data_to_disk_and_clear(activations, weights_data, biases_data, epoch, output_dir):
    """
    This function saves the activation, weights and biases tracking data to disk and then clears the tracking dictionaries to avoid unnecaesary memory usage and potential overflows during longer training runs.

    Args:
        activations (dict): A dictionary containing the activations of each layer in the network
        weights_data (dict): A dictionary containing the weights of each layer in the network
        biases_data (dict): A dictionary containing the biases of each layer in the network
        epoch (int): The current epoch number
        output_dir (str): The path to the directory to save the data to
    """

    if len(activations) != 0:
        activ_path = output_dir + "Activation Data/"
        os.makedirs(activ_path, exist_ok=True)
        # Save the activations to a file named 'activations_epoch_{epoch}.pt'
        torch.save(activations, activ_path + f'activations_epoch_{epoch}.pt')

        # Clear the activations dictionary to free up memory
        activations.clear()

    if len(weights_data) != 0:
        weight_path = output_dir + "Weights Data/"
        os.makedirs(weight_path, exist_ok=True)
        # Save the weights to a file named 'weights_epoch_{epoch}.pt'
        torch.save(weights_data, weight_path + f'weights_epoch_{epoch}.pt')

        # Clear the weights dictionary to free up memory
        weights_data.clear()

    if len(biases_data) != 0:
        bias_path = output_dir + "Biases Data/"
        os.makedirs(bias_path, exist_ok=True)
        # Save the weights to a file named 'weights_epoch_{epoch}.pt'
        torch.save(biases_data, bias_path + f'biases_epoch_{epoch}.pt')

        # Clear the weights dictionary to free up memory
        biases_data.clear()

#%% - Data Gathering Functions
# Tracks confidence of each pixel as histogram per epoch with line showing the detection threshold
def belief_telemetry(data, reconstruction_threshold, epoch, settings, plot_or_save=0):

    """
    This function creates histogram plots of the pixel values recorded by the belief telemtry system [which needs renaming and reworking to simplify. (should be reconstruction thresholding telemtetry?)]  whihc records the values directly out of the netwrok before our reconstruction thresholding in the custom renomalisation is applied. This is important to keep ttrack of what is occusring before our iytput processing as it may be hiding errors.
    The histogram also shows the reconstruction threshold to show the distribution of values above and below the threshold.

    Args:
        data (torch tensor): The image batch. Shape [B, C, H, W]?????????????????????????
        reconstruction_threshold (float): The threshold used in the custom normalisation, used to set the lower limit of the noise point values. Default = 0.5
        epoch (int): The current epoch number
        settings (dict): A dictionary containing the settings for the current training run
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)

    Generates: 
        A histogram plot saved to disk

    Returns:
        above_threshold (int): The number of pixels in the image above the reconstruction threshold
        below_threshold (int): The number of pixels in the image below the reconstruction threshold
    """
    

    data2 = data.flatten()

    #Plots histogram showing the confidence level of each pixel being a signal point
    _, _, bars = plt.hist(data2, 10, histtype='bar')
    plt.axvline(x= reconstruction_threshold, color='red', marker='|', linestyle='dashed', linewidth=2, markersize=12)
    plt.title("Epoch %s" %epoch)
    plt.bar_label(bars, fontsize=10, color='navy') 
    plt.xlabel("Output Values")
    plt.ylabel("Number of Pixels")
    plt.grid(alpha=0.2)
    Out_Label = graphics_dir + f'{model_save_name} - Reconstruction Telemetry Histogram - Epoch {epoch}.png'
    plot_save_choice(plot_or_save, Out_Label)

    above_threshold = (data2 >= reconstruction_threshold).sum()
    below_threshold = (data2 < reconstruction_threshold).sum()
    return (above_threshold, below_threshold)
                
#Combine all performance metrics into simple test script
def quantify_loss_performance(clean_input_batch, noised_target_batch, time_dimension):
    """
    This function compares each image in the batch to its corresponding target image and calculates a range of enhanced performance metrics for each. The results for each metric are then avaeraged over the batch and this average is appended to the corresponding list of average metrics results for each epoch which is tracked eleshwehre in the enhanced performance tracking system

    Args:
        clean_input_batch (torch tensor): The input image batch. Shape [B, C, H, W]
        noised_target_batch (torch tensor): The target image batch. Shape [B, C, H, W]
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values. Default = 100  
    
    Generates:
        Appends the average performance metrics for the batch to the corresponding list of average metrics results for each epoch which is tracked eleshwehre in the enhanced performance tracking system and needs simplifying and condensing!!!

    """
    loss_mse = []
    loss_mae = []
    loss_snr = []
    loss_psnr = []
    loss_ssim = []
    loss_nmi = []
    loss_cc = []
    loss_true_positive_xy = []
    loss_true_positive_tof = []
    loss_false_positive_xy = [] 
    #loss_nrmse = []

    for i in range(len(clean_input_batch)):
        clean_input = clean_input_batch[i][0]
        noised_target = noised_target_batch[i][0]


        ### ADD IN!!
        #loss_nrmse.append(NMSRE(clean_input, noised_target))

        loss_mse.append(MSE(clean_input, noised_target))
        loss_mae.append(MAE(clean_input, noised_target))
        loss_snr.append(SNR(clean_input, noised_target))
        loss_psnr.append(PSNR(clean_input, noised_target, time_dimension))
        loss_ssim.append(SSIM(clean_input, noised_target, time_dimension))
        loss_nmi.append(NomalisedMutualInformation(clean_input, noised_target))
        loss_cc.append(correlation_coeff(clean_input, noised_target))
        percentage_of_true_positive_xy, percentage_of_true_positive_tof, numof_false_positives_xy = compare_images_pixels(clean_input, noised_target)
        loss_true_positive_xy.append(percentage_of_true_positive_xy)
        loss_true_positive_tof.append(percentage_of_true_positive_tof)
        loss_false_positive_xy.append(numof_false_positives_xy)

    #avg_loss_nrmse.append(np.mean(loss_nrmse))
    avg_loss_mse.append(np.mean(loss_mse))
    avg_loss_mae.append(np.mean(loss_mae))
    avg_loss_snr.append(np.mean(loss_snr))
    avg_loss_psnr.append(np.mean(loss_psnr))
    avg_loss_ssim.append(np.mean(loss_ssim))
    avg_loss_nmi.append(np.mean(loss_nmi))
    avg_loss_cc.append(np.mean(loss_cc))
    avg_loss_true_positive_xy.append(np.mean(loss_true_positive_xy))
    avg_loss_true_positive_tof.append(np.mean(loss_true_positive_tof))
    avg_loss_false_positive_xy.append(np.mean(loss_false_positive_xy))

#%% - Data Output Functions

        
def colour_code_excel_file(file_path):
    """

    This function will take an excel file and colour code the cells based on the value of the cell. The colour coding is based on a colour scale from blue to red with blue being the lowest value and red being the highest value. The colour scale is applied to each column individually.

    Args:
        file_path (str): The path to the excel file to be colour coded
    
    Generates:
        Color coding in the excel file which is then saved back to disk in place of the original file
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Define the color scale rule for the conditional formatting
    color_scale = DataBarRule(
        start_type='num',
        start_value=0,
        end_type='num',
        end_value=1,
        color='0000FF',
        showValue=True,
    )

    # Apply the color scale rule to each column
    for column in sheet.columns:
        column_letter = get_column_letter(column[0].column)
        column_range = f'{column_letter}2:{column_letter}{sheet.max_row}'  # Assuming data starts from row 2
        column_cells = list(column)[1:]  # Skip the header cell

        for cell in column_cells:
            cell.number_format = '0.00'  # Optional: Format the cell as desired

        sheet.conditional_formatting.add(column_range, color_scale)

    #resave the workbook to the same file
    workbook.save(file_path)

# Define a function to create a dictionary from the given settings
def create_settings_dict(filename):
    """
    This fucntion will create a dictionary of the settings used in the current training run and save it to disk as a json file.

    Args:
        filename (str): The path to the directory to save the settings dictionary to
    """


    settings_dict = {
        "dataset_title": dataset_title,
        "model_save_name": model_save_name,
        "xdim": xdim,
        "ydim": ydim,
        "time_dimension": time_dimension,
        "num_epochs": num_epochs,
        "batch_size": batch_size,
        "learning_rate": learning_rate,
        "optim_w_decay": optim_w_decay,
        "latent_dim": latent_dim,
        "fc_input_dim": fc_input_dim,
        "dropout_prob": dropout_prob,
        "reconstruction_threshold": reconstruction_threshold,
        "train_test_split_ratio": train_test_split_ratio,
        "val_set_on": val_set_on,
        "val_test_split_ratio": val_test_split_ratio,
        "loss_vs_sparse_img": loss_vs_sparse_img,
        "loss_function_selection": loss_function_selection,
        "zero_weighting": zero_weighting,
        "nonzero_weighting": nonzero_weighting,
        "fullframe_weighting": fullframe_weighting,
        "signal_points" : signal_points,
        "noise_points" : noise_points,
        "x_std_dev" : x_std_dev,
        "y_std_dev" : y_std_dev,
        "tof_std_dev" : tof_std_dev,
        "start_from_pretrained_model" : start_from_pretrained_model,
        "load_pretrained_optimser" : load_pretrained_optimser,
        "pretrained_model_path" : pretrained_model_path,
        "masking_optimised_binary_norm" : masking_optimised_binary_norm,
        "print_every_other" : print_every_other,
        "plot_or_save" : plot_or_save,
        "num_to_plot" : num_to_plot,
        "save_all_raw_plot_data" : save_all_raw_plot_data,
        "double_precision" : double_precision,
        "shuffle_train_data" : shuffle_train_data,
        "timeout_training" : timeout_training,
        "timeout_time" : timeout_time,
        "record_weights" : record_weights,
        "record_biases" : record_biases,
        "record_activity" : record_activity,
        "compress_activations_npz_output" : compress_activations_npz_output,
        "plot_train_loss" : plot_train_loss,
        "plot_validation_loss" : plot_validation_loss,
        "plot_time_loss" : plot_time_loss,
        "plot_detailed_performance_loss" : plot_detailed_performance_loss,
        "plot_live_time_loss" : plot_live_time_loss,
        "plot_live_training_loss" : plot_live_training_loss,
        "comparative_live_loss" : comparative_live_loss,
        "slide_live_plot_size" : slide_live_plot_size,
        "comparative_loss_titles" : comparative_loss_titles,
        "comparative_loss_paths" : comparative_loss_paths,
        "plot_pixel_threshold_telemetry": plot_pixel_threshold_telemetry,
        "plot_pixel_difference": plot_pixel_difference,  # BROKEN
        "plot_latent_generations": plot_latent_generations,
        "plot_higher_dim": plot_higher_dim,
        "plot_Graphwiz": plot_Graphwiz,
        "print_encoder_debug": print_encoder_debug,
        "print_decoder_debug": print_decoder_debug,
        "print_network_summary": print_network_summary,
        "print_partial_training_losses": print_partial_training_losses,
        "debug_noise_function": debug_noise_function,
        "debug_loader_batch": debug_loader_batch,
        "debug_model_exporter": debug_model_exporter,
        "full_dataset_integrity_check": full_dataset_integrity_check,
        "full_dataset_distribution_check": full_dataset_distribution_check,
        "seeding_value": seeding_value,
        "optimise_hyperparameter": optimise_hyperparameter,
        "hyperparam_to_optimise": hyperparam_to_optimise,
        "set_optimisiation_list_manually": set_optimisiation_list_manually,
        "print_validation_results": print_validation_results,
        "plot_training_time": plot_training_time,
        "perf_analysis_num_files": perf_analysis_num_files,
        "perf_analysis_plot": perf_analysis_plot,
        "perf_analysis_dataset_dir": perf_analysis_dataset_dir,
        "debug_hpo_perf_analysis": debug_hpo_perf_analysis
    }

    # Save the settings dictionary to disk
    with open(filename, "w") as f:
        json.dump(settings_dict, f)

    return settings_dict

#%% - Plotting Functions
def loss_plot(x, y, x_label, y_label, title, save_path, plot_or_save):
    """
    Creates a simple line plot of the given data and saves it to disk

    Args:
        x (list): The x axis data
        y (list): The y axis data
        x_label (str): The label for the x axis
        y_label (str): The label for the y axis
        title (str): The title of the plot
        save_path (str): The path to save the plot to
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)

    Generates:
        A simple line plot saved to disk
    """

    plt.plot(x, y)
    plt.title(title)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.grid(alpha=0.2)
    plot_save_choice(plot_or_save, save_path) 

def comparitive_loss_plot(x_list, y_list, legend_label_list, x_label, y_label, title, save_path, plot_or_save):
    """
    Creates a comparative line plot of the given sets of data and saves it to disk

    Args:
        x_list (list): A list of the x axis data
        y_list (list): A list of the y axis data
        legend_label_list (list): A list of the labels for the legend
        x_label (str): The label for the x axis
        y_label (str): The label for the y axis
        title (str): The title of the plot
        save_path (str): The path to save the plot to
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)

    Generates:
        A comparative line plot saved to disk
    """


    for x, y, legend_label in zip(x_list, y_list, legend_label_list):
        plt.plot(x, y, label=legend_label)
    plt.title(title)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.grid(alpha=0.2)
    plt.legend()
    plot_save_choice(plot_or_save, save_path) 

def create_comparison_plot_data(slide_live_plot_size, epoch, max_epoch_reached, comparative_live_loss, comparative_loss_titles, comparative_epoch_times, comparative_history_da, data=history_da['train_loss']):

    """
    This function creates the comparative live loss plot data that updates during training. It takes the data from the current training model and the comparative models and creates the data lists for the live loss plots then calls the comparative_loss_plot fucntion to create the plots.

    Args:
        slide_live_plot_size (int): The number of epochs to show on the live loss plots
        epoch (int): The current epoch number
        max_epoch_reached (int): The maximum epoch number reached by the current training model
        comparative_live_loss (bool): A flag to set if the comparative live loss plots are enabled
        comparative_loss_titles (list): A list of the titles of the comparative models
        comparative_epoch_times (list): A list of the epoch times of the comparative models
        comparative_history_da (list): A list of the history_da dictionaries of the comparative models
        data (list): The data to be plotted on the live loss plots. Default = history_da['train_loss']
    
    Generates:
        A plot of the live loss data vs epoch number, saved to disk
        A plot of the live loss data vs train time elapsed, saved to disk
    """

    ## Slide View
    low_lim = 0
    if slide_live_plot_size > 0:
        low_lim = max(0, epoch - slide_live_plot_size)

    ## Set data for current training model
    x_list_epochs = [range(low_lim, epoch)]
    x_list_time = [epoch_times_list]

    y_list = [data]
    legend_label_list = ["Training loss"]

    ## Add data for comparative models if user desired
    if comparative_live_loss:
        legend_label_list.extend(comparative_loss_titles)

        for loss_dictionary, epoch_t_list in zip(comparative_history_da, comparative_epoch_times):
            epoch_t_list = epoch_t_list[1:]   # NOTE: FIX FOR EPOCH LISTS BEING SAVeD INCLUDING 0 time which breaks the plot by having one more value than the hsitory da train loss. INVESTIGATE THIS ERROR ROOT CAUSE

            #if epoch > len(loss_dictionary['train_loss']):
                #print("WARNING: Epoch is greater than the number of epochs in the comparative model, this will cause an error in the live loss plot")
                #epoch_t_list.append(np.nan) # protection from surpassing the comaprison data during training.
                #loss_dictionary['train_loss'].append(np.nan) # protection from surpassing the comaprison data during training. 
            
            x_list_epochs.append(range(low_lim,len(loss_dictionary['train_loss'])))    
            x_list_time.append(epoch_t_list)
            y_list.append(loss_dictionary['train_loss'])
    
    ## Create plots
    # can have if staments here for protin time and epochs etc
    Out_Label1 = dir + f'{model_save_name} - Live Train loss.png'
    comparitive_loss_plot(x_list_epochs, y_list, legend_label_list, "Epoch number", "Train loss (ACB-MSE)", "Live Training loss", Out_Label1, plot_or_save)
    Out_Label2 = dir + f'{model_save_name} - Live time loss.png'
    comparitive_loss_plot(x_list_time, y_list, legend_label_list, "Time (s)", "Train loss (ACB-MSE)", "Live Time loss", Out_Label2, plot_or_save)



# Helper function to clean up repeated plot save/show code
def plot_save_choice(plot_or_save, output_file_path=none):
    """
    Function used to set the save/display behavior for all figures and graphs genrated by the program.
    
    Args:
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)
        output_file_path (str): The path to save the plot to. Only needed if plot_or_save is set to 1 or 2. 
    """
    if plot_or_save == 0:
        plt.show()
    elif plot_or_save == 1:  
        plt.savefig(output_file_path, format='png')     
        plt.close()
    elif plot_or_save == 2:
        plt.savefig(output_file_path, format='png')   
        plt.show()
    elif plot_or_save == 3:
        plt.close()
    else:
        print("ERROR: Invalid 'plot_or_save' value set for 'plot_save_choice' function call, please set to 0, 1, 2 or 3\n0 prints plots to terminal (blocking till closed).\n1 saves all plots to disk (non-blocking).\n2 prints to terminal and saves to disk (blocking till closed).\n3 will neither save nor show any plots, they will be immidetly closed, usefull for debugging.")
        exit()
    
# Plots the confidence telemetry data
def plot_telemetry(telemetry, true_num_of_signal_points, plot_or_save=0):
    """
    Plots the 'lit pixel count' and 'unlit pixel count' telemetry data over time, compared to true number of signal points for comparison. Used to track the recovery performance.

    Args:
        telemetry (list): A list of the telemetry data to be plotted
        true_num_of_signal_points (int): The true number of signal points in the data set
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 0 prints plots to terminal (blocking till closed), If set to 1 then saves all end of epoch printouts to disk (non-blocking), if set to 2 then saves outputs whilst also printing for user (blocking till closed)

    Generates:
        A plot of the telemetry data, saved to disk or shown depending on program wide 'plot_or_save' setting
    """
    tele = np.array(telemetry)
    plt.plot(tele[:,0],tele[:,1], color='r', label="Points above threshold") #red = num of points above threshold
    plt.plot(tele[:,0],tele[:,2], color='b', label="Points below threshold") #blue = num of points below threshold
    plt.axhline(y=true_num_of_signal_points, color='g', linestyle='dashed', label="True number of signal points")
    plt.title("Telemetry over epochs")
    plt.xlabel("Epoch number")
    plt.ylabel("Number of Signal Points")
    plt.legend()
    Out_Label = graphics_dir + f'{model_save_name} - Reconstruction Telemetry Histogram - Epoch {epoch}.png'
    plot_save_choice(plot_or_save, Out_Label)
            
#%% - Train, Test, Val and Plot Functions
### Training Function
def train_epoch(encoder, decoder, device, dataloader, loss_fn, optimizer, signal_points, noise_points=0, x_std_dev=0, y_std_dev=0, tof_std_dev=0, time_dimension=100, reconstruction_threshold=0.5, print_partial_training_losses=False, masking_optimised_binary_norm=False, loss_vs_sparse_img=False):
    """
    Training loop for a single epoch

    Args:
        encoder (torch model): The encoder model
        decoder (torch model): The decoder model
        device (torch device): The device to run the training on
        dataloader (torch dataloader): The dataloader to iterate over
        loss_fn (torch loss function): The loss function to use
        optimizer (torch optimizer): The optimizer to use
        signal_points (int) or (tuple): The number of signal points to retain in the signal sparsification preprocess. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch
        noise_points (int) or (tuple): The number of noise points to add in the preprocessing. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        x_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the x shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        y_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the y shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        tof_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the ToF shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values amonst other things. Default = 100
        reconstruction_threshold (float): The threshold used in the custom normalisation, also used to set the lower limit of the noise point values. Default = 0.5
        print_partial_training_losses (bool): A flag to set if the partial training losses are printed to terminal. If set to true partial loss is printed to termial whilst training. If set to false the a TQDM progress bar is used to show processing progress. Default = False
        masking_optimised_binary_norm (bool): A flag to set if the masking optimised binary normalisation is used. If set to true the masking optimised binary normalisation is used, if set to false the gaped normalisation is used. Default = False  [EXPLAIN IN MORE DETAIL!!!!!]
        loss_vs_sparse_img (bool): A flag to set if the loss is calculated against the sparse image or the clean image. If set to true the loss is calculated against the sparse image, if set to false the loss is calculated against the clean image. Default = False

    Returns:
        loss_total/batches (float): The average loss over the epoch calulated by dividing the total sum of loss by the number of batches !! EXPLAIN THIS SIMPLER !!! 

    """
    encoder.train()   
    decoder.train()   

    loss_total = 0.0
    batches = 0

    if print_partial_training_losses:  # Prints partial train losses per batch
        image_loop  = (dataloader)     # No progress bar for the batches
    else:                              # Rather than print partial train losses per batch, instead create progress bar
        image_loop  = tqdm(dataloader, desc='Batches', leave=False) # Creates a progress bar for the batches

    # Iterate the dataloader (we do not need the label values, this is unsupervised learning)
    for image_batch, _ in image_loop: # with "_" we just ignore the labels (the second element of the dataloader tuple

        #Select settings randomly from user ranges
        signal_settings = input_range_to_random_value(signal_points, x_std_dev, y_std_dev, tof_std_dev, noise_points)

        # DATA PREPROCESSING
        with torch.no_grad(): # No need to track the gradients
            sparse_output_batch, sparse_and_resolution_limited_batch, noised_sparse_reslimited_batch = signal_degredation(signal_settings, image_batch, physical_scale_parameters)
            
            if masking_optimised_binary_norm:
                normalised_batch = mask_optimised_normalisation(noised_sparse_reslimited_batch)
                norm_sparse_output_batch = mask_optimised_normalisation(sparse_output_batch)
                normalised_inputs = mask_optimised_normalisation(image_batch)
            else:
                normalised_batch = gaped_normalisation(noised_sparse_reslimited_batch, reconstruction_threshold, time_dimension)
                norm_sparse_output_batch = gaped_normalisation(sparse_output_batch, reconstruction_threshold, time_dimension)
                normalised_inputs = gaped_normalisation(image_batch, reconstruction_threshold, time_dimension)
            
        # Move tensor to the proper device
        image_clean = normalised_inputs.to(device) # Move the clean image batch to the device
        image_sparse = norm_sparse_output_batch.to(device) # Move the sparse image batch to the device
        image_noisy = normalised_batch.to(device) # Move the noised image batch to the device
        
        # Encode data
        encoded_data = encoder(image_noisy) # Encode the noised image batch
        # Decode data
        decoded_data = decoder(encoded_data) # Decode the encoded image batch
        
        if loss_vs_sparse_img:
            loss_comparator = image_sparse
        else:
            loss_comparator = image_clean

        # Evaluate loss
        #if renorm_for_loss_calc
            #decoded_data = gaped_renormalisation_torch(decoded_data, reconstruction_threshold, time_dimension)
            #loss_comparator = gaped_renormalisation_torch(loss_comparator, reconstruction_threshold, time_dimension)
        loss = loss_fn(decoded_data, loss_comparator)  # Compute the loss between the decoded image batch and the clean image batch
        
        # Backward pass
        optimizer.zero_grad() # Reset the gradients
        loss.backward() # Compute the gradients


        if use_tensorboard:
            # Add the gradient values to Tensorboard
            for name, param in encoder.named_parameters():
                writer.add_histogram(name + '/grad', param.grad, global_step=epoch)

            for name, param in decoder.named_parameters():
                writer.add_histogram(name + '/grad', param.grad, global_step=epoch)

        optimizer.step() # Update the parameters
        batches = batches + 1
        loss_total = loss_total + loss.detach().cpu().numpy()

        if print_partial_training_losses:         # Prints partial train losses per batch
            print('\t partial train loss (single batch): %f' % (loss.data))  # Print batch loss value
    
    return loss_total/batches

### Testing Function
def test_epoch(encoder, decoder, device, dataloader, loss_fn, signal_points, noise_points=0, x_std_dev=0, y_std_dev=0, tof_std_dev=0, time_dimension=100, reconstruction_threshold=0.5, print_partial_training_losses=False, masking_optimised_binary_norm=False, loss_vs_sparse_img=False):
    """
    Testing (Evaluation) loop for a single epoch. This function is identical to the training loop except that it does not perform the backward pass and parameter update steps and the model is run in eval mode. Additionaly the dataset used is the test dataset rather than the training dataset so that the data is unseen by the model.

    Args:

        encoder (torch model): The encoder model
        decoder (torch model): The decoder model
        device (torch device): The device to run the training on
        dataloader (torch dataloader): The dataloader to iterate over
        loss_fn (torch loss function): The loss function to use
        signal_points (int) or (tuple): The number of signal points to retain in the signal sparsification preprocess. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch
        noise_points (int) or (tuple): The number of noise points to add in the preprocessing. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        x_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the x shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        y_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the y shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        tof_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the ToF shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values amonst other things. Default = 100
        reconstruction_threshold (float): The threshold used in the custom normalisation, also used to set the lower limit of the noise point values. Default = 0.5
        print_partial_training_losses (bool): A flag to set if the partial training losses are printed to terminal. If set to true partial loss is printed to termial whilst training. If set to false the a TQDM progress bar is used to show processing progress. Default = False
        masking_optimised_binary_norm (bool): A flag to set if the masking optimised binary normalisation is used. If set to true the masking optimised binary normalisation is used, if set to false the gaped normalisation is used. Default = False  [EXPLAIN IN MORE DETAIL!!!!!]
        loss_vs_sparse_img (bool): A flag to set if the loss is calculated against the sparse image or the clean image. If set to true the loss is calculated against the sparse image, if set to false the loss is calculated against the clean image. Default = False

    Returns:
        loss_total/batches (float): The average loss over the epoch calulated by dividing the total sum of loss by the number of batches !! EXPLAIN THIS SIMPLER !!!

    """
    # Set evaluation mode for encoder and decoder
    encoder.eval() # Evaluation mode for the encoder
    decoder.eval() # Evaluation mode for the decoder
    with torch.no_grad(): # No need to track the gradients

        loss_total = 0.0
        batches = 0

        if print_partial_training_losses:  # Prints partial train losses per batch
            image_loop = (dataloader)     # No progress bar for the batches
        else:                              # Rather than print partial train losses per batch, instead create progress bar
            image_loop = tqdm(dataloader, desc='Testing', leave=False, colour="yellow") # Creates a progress bar for the batches

        for image_batch, _ in image_loop:  

            #Select settings randomly from user range
            signal_settings = input_range_to_random_value(signal_points, x_std_dev, y_std_dev, tof_std_dev, noise_points)
            sparse_output_batch, sparse_and_resolution_limited_batch, noised_sparse_reslimited_batch = signal_degredation(signal_settings, image_batch, physical_scale_parameters)

            if masking_optimised_binary_norm:
                normalised_batch = mask_optimised_normalisation(noised_sparse_reslimited_batch)
                image_batch_norm = mask_optimised_normalisation(image_batch)
            else:
                normalised_batch = gaped_normalisation(noised_sparse_reslimited_batch, reconstruction_threshold, time_dimension)
                image_batch_norm = gaped_normalisation(image_batch, reconstruction_threshold, time_dimension)
            
            image_clean = image_batch_norm.to(device) # Move the clean image batch to the device
            image_noisy = normalised_batch.to(device) # Move the noised image batch to the device

            # Encode data
            encoded_data = encoder(image_noisy) # Encode the noised image batch
            # Decode data
            decoded_data = decoder(encoded_data) # Decode the encoded image batch

            if loss_vs_sparse_img:
                loss_comparator = sparse_output_batch
            else:
                loss_comparator = image_clean

            # Evaluate loss
            loss = loss_fn(decoded_data, loss_comparator)  # Compute the loss between the decoded image batch and the clean image batch
            
            batches = batches + 1
            loss_total = loss_total + loss.detach().cpu().numpy()

            #Run additional perfomrnace metric loss functions for final plots, this needs cleaning up!!!!!
            quantify_loss_performance(loss_comparator, decoded_data, time_dimension)

    return loss_total/batches

### Validation Function
def validation_routine(encoder, decoder, device, dataloader, loss_fn, signal_points, noise_points=0, x_std_dev=0, y_std_dev=0, tof_std_dev=0, time_dimension=100, reconstruction_threshold=0.5, print_partial_training_losses=False, masking_optimised_binary_norm=False, loss_vs_sparse_img=False):
    """
    Validation loop for a single epoch. This function is identical to the test/evaluation loop except that it is used for hyperparamter evaluation to evaluate between differnt models. This function is not used during training, only for hyperparameter evaluation. Again it uses a previosuly unseen dataset howevr this one is fixed and not randomly selected from the dataset so as to provide a fixed point of reference for direct model comparison.
    
    Args:
        encoder (torch model): The encoder model
        decoder (torch model): The decoder model
        device (torch device): The device to run the training on
        dataloader (torch dataloader): The dataloader to iterate over
        loss_fn (torch loss function): The loss function to use
        signal_points (int) or (tuple): The number of signal points to retain in the signal sparsification preprocess. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch
        noise_points (int) or (tuple): The number of noise points to add in the preprocessing. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        x_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the x shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        y_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the y shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        tof_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the ToF shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values amonst other things. Default = 100
        reconstruction_threshold (float): The threshold used in the custom normalisation, also used to set the lower limit of the noise point values. Default = 0.5
        print_partial_training_losses (bool): A flag to set if the partial training losses are printed to terminal. If set to true partial loss is printed to termial whilst training. If set to false the a TQDM progress bar is used to show processing progress. Default = False
        masking_optimised_binary_norm (bool): A flag to set if the masking optimised binary normalisation is used. If set to true the masking optimised binary normalisation is used, if set to false the gaped normalisation is used. Default = False  [EXPLAIN IN MORE DETAIL!!!!!]
        loss_vs_sparse_img (bool): A flag to set if the loss is calculated against the sparse image or the clean image. If set to true the loss is calculated against the sparse image, if set to false the loss is calculated against the clean image. Default = False

    Returns:
        loss_total/batches (float): The average loss over the epoch calulated by dividing the total sum of loss by the number of batches !! EXPLAIN THIS SIMPLER !!!

    """
    # Set evaluation mode for encoder and decoder
    encoder.eval() # Evaluation mode for the encoder
    decoder.eval() # Evaluation mode for the decoder
    with torch.no_grad(): # No need to track the gradients

        loss_total = 0.0
        batches = 0

        if print_partial_training_losses:  # Prints partial train losses per batch
            image_loop = (dataloader)     # No progress bar for the batches
        else:                              # Rather than print partial train losses per batch, instead create progress bar
            image_loop = tqdm(dataloader, desc='Validation', leave=False, colour="green") # Creates a progress bar for the batches

        for image_batch, _ in image_loop:  

            #Select settings randomly from user range
            signal_settings = input_range_to_random_value(signal_points, x_std_dev, y_std_dev, tof_std_dev, noise_points)
            sparse_output_batch, sparse_and_resolution_limited_batch, noised_sparse_reslimited_batch = signal_degredation(signal_settings, image_batch, physical_scale_parameters)

            if masking_optimised_binary_norm:
                normalised_batch = mask_optimised_normalisation(noised_sparse_reslimited_batch)
                image_batch_norm = mask_optimised_normalisation(image_batch)
            else:
                normalised_batch = gaped_normalisation(noised_sparse_reslimited_batch, reconstruction_threshold, time_dimension)
                image_batch_norm = gaped_normalisation(image_batch, reconstruction_threshold, time_dimension)
            
            image_clean = image_batch_norm.to(device) # Move the clean image batch to the device
            image_noisy = normalised_batch.to(device) # Move the noised image batch to the device

            # Encode data
            encoded_data = encoder(image_noisy) # Encode the noised image batch
            # Decode data
            decoded_data = decoder(encoded_data) # Decode the encoded image batch

            if loss_vs_sparse_img:
                loss_comparator = sparse_output_batch
            else:
                loss_comparator = image_clean

            # Evaluate loss
            loss = loss_fn(decoded_data, loss_comparator)  # Compute the loss between the decoded image batch and the clean image batch
            
            batches = batches + 1
            loss_total = loss_total + loss.detach().cpu().numpy()

            #Run additional perfomrnace metric loss functions for final plots, this needs cleaning up!!!!!
            #quantify_loss_performance(loss_comparator, decoded_data, time_dimension)
    
    return loss_total/batches

### Plotting function
def plot_epoch_data(encoder, decoder, epoch, model_save_name, time_dimension, reconstruction_threshold, signal_points, noise_points=0, x_std_dev=0, y_std_dev=0, tof_std_dev=0, physical_scale_parameters=[1,1,1], n=10):       #Defines a function for plotting the output of the autoencoder. And also the input + clean training data? Function takes inputs, 'encoder' and 'decoder' which are expected to be classes (defining the encode and decode nets), 'n' which is the number of ?????Images in the batch????, and 'noise_factor' which is a multiplier for the magnitude of the added noise allowing it to be scaled in intensity.  
    
    """
    Plots the output of the autoencoder in a variety of ways to track its perfromance and abilities during the training cycle. 

    Args:
        encoder (torch model): The encoder model
        decoder (torch model): The decoder model
        epoch (int): The current epoch number
        model_save_name (str): The name of the model being trained
        time_dimension (int): The number of time steps in the data set, used to set the upper limit of the noise point values amonst other things. Default = 100
        reconstruction_threshold (float): The threshold used in the custom normalisation, also used to set the lower limit of the noise point values. Default = 0.5
        signal_points (int) or (tuple): The number of signal points to retain in the signal sparsification preprocess. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch
        noise_points (int) or (tuple): The number of noise points to add in the preprocessing. If given as int then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        x_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the x shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        y_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the y shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        tof_std_dev (float) or (tuple): The standard deviation of the gaussian distribution to draw the ToF shift from. If given as float then number will be constant over training, if given as tuple then a random value will be selected from the range given for each batch. Default = 0
        physical_scale_parameters (list): A list of the physical scale parameters for the data set. Default = [1,1,1]
        n is the number of images to plot in the 2d comparison. Default = 10

    Generates:
        A variety of plots showing the output of the autoencoder
    
    """
    encoder.eval()                                   #.eval() is a kind of switch for some specific layers/parts of the model that behave differently during training and inference (evaluating) time. For example, Dropouts Layers, BatchNorm Layers etc. You need to turn off them during model evaluation, and .eval() will do it for you. In addition, the common practice for evaluating/validation is using torch.no_grad() in pair with model.eval() to turn off gradients computation
    decoder.eval()                                   #Simarlary as above

    encoder.to(device)
    decoder.to(device)
    # Unpack physical scale parameters
    x_scale, y_scale, tof_scale = physical_scale_parameters

    # Initialise lists for true and recovered signal point values 
    number_of_true_signal_points = []
    number_of_recovered_signal_points = []

    # 2D Input/Output Comparison Plots 
    plt.figure(figsize=(16,9))                                      #Sets the figure size
    loop = tqdm(range(n), desc='Plotting 2D Comparisons', leave=False, colour="green") 
    for i in loop:                                                #Runs for loop where 'i' itterates over 'n' total values which range from 0 to n-1

        #Select settings randomly from user range
        signal_settings = input_range_to_random_value(signal_points, x_std_dev, y_std_dev, tof_std_dev, noise_points)

        # Load input image
        img = train_dataset[i][0].unsqueeze(0) # Unsqueeze to add artificial batch dimension
        
        # Save the number of true signal points on the input image 
        number_of_true_signal_points.append(signal_settings[0])  
        
        #global image_noisy   
              
        # Create degraded images                     
        sparse_output_batch, sparse_and_resolution_limited_batch, noised_sparse_reslimited_batch = signal_degredation(signal_settings, img, physical_scale_parameters)

        # Normalise the noised image
        if masking_optimised_binary_norm:
            normalised_batch = mask_optimised_normalisation(noised_sparse_reslimited_batch)
        else:
            normalised_batch = gaped_normalisation(noised_sparse_reslimited_batch, reconstruction_threshold, time_dimension)
        
        # Run the autoencoder on the noised data  
        with torch.no_grad():  
            normalised_batch = normalised_batch.to(device)                                       
            rec_img = decoder(encoder(normalised_batch))   

        #Determine the number of signal points on the recovered image 
        int_rec_sig_points = (rec_img >= reconstruction_threshold).sum()      
        number_of_recovered_signal_points.append(int(int_rec_sig_points.cpu().numpy()))

        test_image = img.cpu().squeeze()
        sparse_im = sparse_output_batch.cpu().squeeze()
        reslim_im = sparse_and_resolution_limited_batch.cpu().squeeze()
        network_input_image = normalised_batch.cpu().squeeze() 
        recovered_test_image = rec_img.cpu().squeeze().numpy()

        #clean up lines
        in_im = test_image   
        noise_im = gaped_renormalisation(network_input_image, reconstruction_threshold, time_dimension)
        rec_im = gaped_renormalisation(recovered_test_image, reconstruction_threshold, time_dimension)
        masked_im = masking_recovery(noise_im, rec_im, time_dimension)

        #cmap = cm.get_cmap('viridis')
        #cmap.set_under('k') # set the color for 0 to black ('k')

        #Following section generates the img plots for the original(labels), noised, and denoised data)
        ax = plt.subplot(6,n,i+1)                                       #Creates a number of subplots for the 'Original images??????' i.e the labels. the position of the subplot is i+1 as it falls in the first row
        plt.imshow(in_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)           #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('EPOCH %s \nOriginal images' %(epoch))               #When above condition is reached, the plots title is set                                   #When above condition is reached, the plots title is set

        ax = plt.subplot(6, n, i + 1 + n)                                   #Creates a number of subplots for the 'Corrupted images??????' i.e the labels. the position of the subplot is i+1+n as it falls in the second row
        plt.imshow(sparse_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)   #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('Sparsity Applied')                                  #When above condition is reached, the plots title is set

        ax = plt.subplot(6, n, i + 1 + n + n)                               #Creates a number of subplots for the 'Reconstructed images??????' i.e the labels. the position of the subplot is i+1+n+n as it falls in the third row
        plt.imshow(reslim_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)       #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('Resoloution Limited')                             #When above condition is reached, the plots title is set 

        ax = plt.subplot(6, n, i + 1 + n + n + n)                                   #Creates a number of subplots for the 'Corrupted images??????' i.e the labels. the position of the subplot is i+1+n as it falls in the second row
        plt.imshow(noise_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)   #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('Corrupted images')                                  #When above condition is reached, the plots title is set

        ax = plt.subplot(6, n, i + 1 + n + n + n + n)                               #Creates a number of subplots for the 'Reconstructed images??????' i.e the labels. the position of the subplot is i+1+n+n as it falls in the third row
        plt.imshow(rec_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)       #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('Reconstructed images')                             #When above condition is reached, the plots title is set 

        ax = plt.subplot(6, n, i + 1 + n + n + n + n + n)                               #Creates a number of subplots for the 'Reconstructed images??????' i.e the labels. the position of the subplot is i+1+n+n as it falls in the third row
        plt.imshow(masked_im.T, cmap='gist_gray', vmin=0, vmax=time_dimension)       #plt.imshow plots an image. The arguments for imshow are, 'image data array' and cmap= which is the colour map. #.squeeze() acts on a tensor and returns a tensor, it removes all dimensions of the tensor that are of length 1, (A×1×B) becomes (AxB) where A and B are greater than 1 #.numpy() creates a numpy array from a tensor #!!! is the .cpu part becuase the code was not made to accept the gpu/cpu check i made????
        ax.get_xaxis().set_visible(False)                                   #Hides the x axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        ax.get_yaxis().set_visible(False)                                   #Hides the y axis from showing in the plot as we are plotting images not graphs (we may want to retain axis?)
        if i == n//2:                                                       #n//2 divides n by 2 without any remainder, i.e 6//2=3 and 7//2=3. So this line checks to see if i is equal to half of n without remainder. it will be yes once in the loop. not sure of its use
            ax.set_title('Masked Reconstruction')                             #When above condition is reached, the plots title is set 

    plt.subplots_adjust(left=0.01,              #Adjusts the exact layout of the plots including whwite space round edges
                    bottom=0.02, 
                    right=0.99, 
                    top=0.95, 
                    hspace=0.3
                    )     

    Out_Label = graphics_dir + f'{model_save_name} - Epoch {epoch}.png' #creates the name of the file to be saved
    plot_save_choice(plot_or_save, Out_Label) #saves the plot if plot_or_save is set to 1, if 0 it displays, if 2 it displays and saves
    plt.close()

    # 3D Reconstruction
    in_im, sparse_im, reslim_im, noise_im, rec_im, masked_im = reconstruct_3D(in_im, sparse_im, reslim_im, noise_im, rec_im, masked_im) #reconstructs the 3D image using the reconstruct_3D function

    # 3D Plottting
    if rec_im.ndim != 1:                       # Checks if there are actually values in the reconstructed image, if not no image is aseved/plotted
        fig, axs = plt.subplots(2, 3, figsize=(32,20), subplot_kw={'projection': '3d'})
        ax1, ax2, ax3, ax4, ax5, ax6 = axs.flatten()
        fig.suptitle(f"3D Reconstruction - Epoch {epoch}") #sets the title of the plot

        ax1.set_title("Input Image") #sets the title of the plot
        ax1.scatter(in_im[:,0], in_im[:,1], in_im[:,2]) #plots the 3D scatter plot for input 

        ax2.set_title("Sparse Image") #sets the title of the plot
        ax2.scatter(sparse_im[:,0], sparse_im[:,1], sparse_im[:,2]) #plots the 3D scatter plot for sparse image
    
        ax3.set_title("Resolution Limited Image") #sets the title of the plot
        ax3.scatter(reslim_im[:,0], reslim_im[:,1], reslim_im[:,2]) #plots the 3D scatter plot for reslim image

        ax4.set_title("Noised Image") #sets the title of the plot
        ax4.scatter(noise_im[:,0], noise_im[:,1], noise_im[:,2]) #plots the 3D scatter plot for noised image

        ax5.set_title("Reconstructed Image") #sets the title of the plot
        ax5.scatter(rec_im[:,0], rec_im[:,1], rec_im[:,2]) #plots the 3D scatter plot for reconstructed image

        try: ## NOTE THIS ERROR NEEDS FIXING. IT IS CAUSED BY A SITUATION WHERE DATA DOES COME BACK THAT IS GREATER THAN THE RECON CUTTOFF SO 3D PLOTS ARE GENERATED HOWVER NO POINTS LIE IN CORRECT PLACE FOR MASKING SO THE MASK CONTAINS NOTHING. THEN THE MASK WILL BE WRONG DIMS AND CASUE THE PLOT ERROR HERE. FIX
            ax6.set_title("Masked Reconstructed Image") #sets the title of the plot
            ax6.scatter(masked_im[:,0], masked_im[:,1], masked_im[:,2]) #plots the 3D scatter plot for masked image
        except:
            print("ERROR OCCURED IN MASKING 3D PLOT! INVESTIGATE!")


        for ax in [ax1, ax2, ax3, ax4, ax5, ax6]:
            ax.set_zlim(0, time_dimension)
            ax.set_xlim(0, 128)
            ax.set_ylim(0, 88)
        
            if use_physical_values_for_plot_axis:
                # Set axis labels
                ax.set_xlabel('x (mm)')
                ax.set_ylabel('y (mm)')
                ax.set_zlabel('time (ns)')
                # Apply tick format conversion for x, y, and z axes from 'pixels' to physical values
                ax.xaxis.set_major_formatter(FuncFormatter(lambda x_scale, tick_number: tick_number * x_scale))
                ax.yaxis.set_major_formatter(FuncFormatter(lambda y_scale, tick_number: tick_number * y_scale))
                ax.zaxis.set_major_formatter(FuncFormatter(lambda time_scale, tick_number: tick_number * time_scale))
            
            else:
                # Set axis labels
                ax.set_xlabel('x (pixels)')
                ax.set_ylabel('y (pixels)')
                ax.set_zlabel('time (pixels)')

        fig.tight_layout()

        Out_Label = graphics_dir + f'{model_save_name} 3D Reconstruction - Epoch {epoch}.png' #creates the name of the file to be saved
        plot_save_choice(plot_or_save, Out_Label) #saves the plot if plot_or_save is set to 1, if 0 it displays, if 2 it displays and saves

        if plot_pixel_threshold_telemetry == 1:      #if plot_pixel_threshold_telemetry is set to 1, then the telemetry plots are generated
            above_threshold, below_threshold = belief_telemetry(recovered_test_image, reconstruction_threshold, epoch+1, settings, plot_or_save)    #calls the belief_telemetry function to generate the telemetry plots
            telemetry.append([epoch, above_threshold, below_threshold]) #appends the telemetry data to the telemetry list

    return(number_of_true_signal_points, number_of_recovered_signal_points, in_im, noise_im, rec_im)        #returns the number of true signal points, number of recovered signal points, input image, noised image and reconstructed image 

#%% - Custom Loss Fn Classes


#from Loss_Functions.ada_weighted_custom_split_loss import ada_weighted_custom_split_loss  #NOT YET MOVED, NEEDS DEPEMDECIOES IN THIS CODE SORTED FIRST

# Weighted Custom Split Loss Function
def ada_weighted_custom_split_loss(reconstructed_image, target_image, zero_weighting=zero_weighting, nonzero_weighting=nonzero_weighting):
    """
    Calculates the weighted error loss between target_image and reconstructed_image.
    The loss for zero pixels in the target_image is weighted by zero_weighting, and the loss for non-zero
    pixels is weighted by nonzero_weighting and both have loss functions as passed in by user.

    Args:
    - target_image: a tensor of shape (B, C, H, W) containing the target image
    - reconstructed_image: a tensor of shape (B, C, H, W) containing the reconstructed image
    - zero_weighting: a scalar weighting coefficient for the MSE loss of zero pixels
    - nonzero_weighting: a scalar weighting coefficient for the MSE loss of non-zero pixels

    Returns:
    - weighted_mse_loss: a scalar tensor containing the weighted MSE loss
    """
    
    # Get the indices of 0 and non 0 values in target_image as a mask for speed
    zero_mask = (target_image == 0)
    nonzero_mask = ~zero_mask         # Invert mask
    
    # Get the values in target_image
    values_zero = target_image[zero_mask]
    values_nonzero = target_image[nonzero_mask]
    
    # Get the corresponding values in reconstructed_image
    corresponding_values_zero = reconstructed_image[zero_mask]
    corresponding_values_nonzero = reconstructed_image[nonzero_mask]
    
    # Get the loss functions
    loss_func_zeros = split_loss_functions[0]
    loss_func_nonzeros = split_loss_functions[1]
    
    # Compute the MSE losses
    zero_loss = loss_func_zeros(corresponding_values_zero, values_zero)
    nonzero_loss = loss_func_nonzeros(corresponding_values_nonzero, values_nonzero)

    # Protection from there being no 0 vals or no non zero vals, which then retunrs nan for MSE and creates a nan overall MSE return (which is error)
    if torch.isnan(zero_loss):
        zero_loss = 0
    if torch.isnan(nonzero_loss):
        nonzero_loss = 0
    
    # Sum losses with weighting coefficiants 
    weighted_mse_loss = (zero_weighting * zero_loss) + (nonzero_weighting * nonzero_loss) 
    
    return weighted_mse_loss



#%% - Program begins
print("\n \nProgram Initalised - Welcome to DC3D Trainer\n")  #prints the welcome message

# Following section checks if a CUDA enabled GPU is available. If found it is selected as the 'device' to perform the tensor opperations. If no CUDA GPU is found the 'device' is set to CPU (much slower) 
device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
print(f'Selected compute device: {device}\n')  #Informs user if running on CPU or GPU


#%% - # Hyperparameter Optimiser
if optimise_hyperparameter:
    HTO_values = set_optimisiation_list_manually
    print(f"Hyperparameter optimisation is enabled, the following hyperparameters will be optimised: {hyperparam_to_optimise}")
    val_loop_range = HTO_values #val_loop_range is the number of values in the HTO_values list, i.e. the number of times the model will be trained and evaluated
else: 
    val_loop_range = range(1,2,1)

for HTO_val in val_loop_range: #val_loop is the number of times the model will be trained and evaluated
    if optimise_hyperparameter:
        globals()[hyperparam_to_optimise] = HTO_val
        print(f"Current Test: {hyperparam_to_optimise} value = {HTO_val}\n")





    #%% - # Input / Output Path Initialisation
    if optimise_hyperparameter:
        results_output_path = results_output_path_1 + model_save_name + "_" + hyperparam_to_optimise + f" Optimisation\\{HTO_val}\\"

    # Create output directory if it doesn't exist
    dir = results_output_path + model_save_name + " - Training Results/"
    os.makedirs(dir, exist_ok=True)

    # Create output directory for images if it doesn't exist
    graphics_dir = dir + "Output_Graphics/"
    os.makedirs(graphics_dir, exist_ok=True)

    raw_plotdata_output_dir = dir + "Raw_Data_Output/"
    os.makedirs(raw_plotdata_output_dir, exist_ok=True)

    model_output_dir = dir + "Model_Deployment/"
    os.makedirs(model_output_dir, exist_ok=True)

    # Joins up the parts of the differnt output files save paths
    full_model_path = model_output_dir + model_save_name + " - Model.pth"
    full_statedict_path = model_output_dir + model_save_name + " - Model + Optimiser State Dicts.pth"

    full_activity_filepath = dir + model_save_name + " - Activity.npz"
    full_netsum_filepath = dir + model_save_name + " - Network Summary.txt"


    # Joins up the parts of the differnt input dataset load paths
    train_dir = data_path + dataset_title

    #%% - Bodge clean up 
    ###CLEAN UP THIS METHOD TO SOMTHING BETTER!!!!!!
    epoch_avg_loss_mse = []
    epoch_avg_loss_mae = []
    epoch_avg_loss_snr = []
    epoch_avg_loss_psnr = []
    epoch_avg_loss_ssim = []
    epoch_avg_loss_nmi = []
    epoch_avg_loss_cc = []
    epoch_avg_loss_true_positive_xy = []
    epoch_avg_loss_true_positive_tof = []
    epoch_avg_loss_false_positive_xy = []

    # Initialises pixel belief telemetry
    telemetry = [[0,0.5,0.5]]                # Initalises the telemetry memory, starting values are 0, 0.5, 0.5 which corrspond to epoch(0), above_threshold(0.5), below_threshold(0.5)
    
    # Create a summary writer
    writer = SummaryWriter()


    #%% - Initialises seeding values to RNGs
    if seeding_value:
        torch.manual_seed(seeding_value)
        np.random.seed(seeding_value)
        print("Seeding with value:", seeding_value)
    else:
        torch.seed()                                    # Set the seed for the RNGs
        np.random.seed()

    T_seed = torch.initial_seed()
    N_seed = np.random.get_state()[1][0]


    #%% - Set loss function choice

    # List of all availible loss functions
    availible_loss_functions = [ACBLoss(zero_weighting, nonzero_weighting), 
                                ffACBLoss(zero_weighting, nonzero_weighting, fullframe_weighting), 
                                True3DLoss(zero_weighting=1, nonzero_weighting=1, timesteps=1000), 
                                boostedffACBLoss(zero_weighting, nonzero_weighting, fullframe_weighting, ff_loss, boost=BOOST),                                
                                simple3Dloss(zero_weighting=1, nonzero_weighting=1, virtual_t_weighting=1, virtual_x_weighting=1, virtual_y_weighting=1, timesteps=1000), 
                                ACBLoss3D(zero_weighting=1, nonzero_weighting=1, virtual_t_weighting=1, virtual_x_weighting=None, virtual_y_weighting=None, timesteps=1000), 
                                torch.nn.MSELoss(), 
                                torch.nn.BCELoss(), 
                                torch.nn.L1Loss(), 
                                ada_SSE_loss, 
                                ada_weighted_custom_split_loss, 
                                WeightedPerfectRecoveryLoss(),
                                NEWESTACB3dloss2024(batch_size, ydim, xdim)]    
   
    # List of all availible loss function labels
    loss_function_labels = ["ACB_MSE", 
                            "ffACB_MSE", 
                            "True3DLoss",
                            "boostedffACB_MSE",
                            "simple3Dloss",
                            "ACBLoss3D", 
                            "boostedffACB_MSE",
                            "MSE", 
                            "BCE", 
                            "MAE", 
                            "SSE", 
                            "Split Loss", 
                            "Weighted Perfect Recovery",
                            "2024 - 3D LOSS"]
    
    #loss_fn = 
    loss_fn = availible_loss_functions[loss_function_selection]            # Sets loss function based on user input of parameter loss_function_selection
    loss_fn_label = loss_function_labels[loss_function_selection]          # Sets loss function label based on user input of parameter loss_function_selection

    # Set loss function choice for split loss (if loss function choice is set to ada weighted custom split loss)
    availible_split_loss_functions = [torch.nn.MSELoss(), torch.nn.BCELoss(), torch.nn.L1Loss(), ada_SSE_loss]    # List of all availible loss functions is set to ada_weighted_custom_split_loss
    split_loss_functions = [availible_split_loss_functions[zeros_loss_choice], availible_split_loss_functions[nonzero_loss_choice]] # Sets loss functions based on user input


    #%% - Create record of all user input settings, to add to output data for testing and keeping track of settings
    settings = {}  # Creates empty dictionary to store settings 
    settings["Epochs"] = num_epochs     # Adds the number of epochs to the settings dictionary
    settings["Batch Size"] = batch_size # Adds the batch size to the settings dictionary
    settings["Learning Rate"] = learning_rate # Adds the learning rate to the settings dictionary
    settings["Optimiser Decay"] = optim_w_decay # Adds the optimiser decay to the settings dictionary
    settings["Dropout Probability"] = dropout_prob # Adds the dropout probability to the settings dictionary
    settings["Latent Dimension"] = latent_dim # Adds the latent dimension to the settings dictionary
    settings["Noise Points"] = noise_points # Adds the noise points to the settings dictionary
    settings["Train Split"] = train_test_split_ratio # Adds the train split to the settings dictionary
    settings["Val/Test Split"] = val_test_split_ratio   # Adds the val/test split to the settings dictionary
    settings["Dataset"] = dataset_title # Adds the dataset title to the settings dictionary
    settings["Time Dimension"] = time_dimension # Adds the time dimension to the settings dictionary
    settings["Seed Val"] = seeding_value # Adds the seed value to the settings dictionary
    settings["Reconstruction Threshold"] = reconstruction_threshold # Adds the reconstruction threshold to the settings dictionary
    settings["Double Precision"] = double_precision
    settings["shuffle_train_data"] = shuffle_train_data
    settings["record_weights"] = record_weights
    settings["record_biases"] = record_biases
    settings["record_activity"] = record_activity
    settings["compress_activations_npz_output"] = compress_activations_npz_output
    settings["timeout_training"] = timeout_training
    settings["timeout_time"] = timeout_time
 




    #%% - Dataset Pre-tests
    # Dataset Integrity Check   
    if full_dataset_integrity_check: #if full_dataset_integrity_check is set to True, then the scan type is set to full (slow)
        scantype = "full"
    else:
        scantype = "quick"

    # TERMINAL PRINT: Report test type and result
    print(f"Testing training dataset integrity, with {scantype} scan")  #prints the scan type
    
    # CHECK: Dataset Distribution Check
    dataset_integrity_check(train_dir, full_test=full_dataset_integrity_check, print_output=True) #checks the integrity of the training dataset
    #print("Test completed\n")

    # CHECK: Dataset Distribution Check
    if full_dataset_distribution_check: #if full_dataset_distribution_check is set to True, then the dataset distribution is checked
        print("\nTesting training dataset signal distribution")
        dataset_distribution_tester(train_dir, time_dimension, ignore_zero_vals_on_plot=True, output_image_dir=graphics_dir) #checks the distribution of the training dataset
        #print("Test completed\n")

    # CHECK: Num of images in path greater than batch size choice? 
    num_of_files_in_path = len(os.listdir(data_path + dataset_title + '/Data/')) #number of files in path
    if num_of_files_in_path < batch_size: #if the number of files in the path is less than the batch size, user is promted to input a new batch size
        print("Error, the path selected has", num_of_files_in_path, "image files, which is", (batch_size - num_of_files_in_path) , "less than the chosen batch size. Please select a batch size less than the total number of images in the directory")
        batch_err_message = "Choose new batch size, must be less than total amount of images in directory", (num_of_files_in_path) #creates the error message
        batch_size = int(input(batch_err_message))  #!!! not sure why input message is printing with wierd brakets and speech marks in the terminal? Investigate

    # TERMINAL PRINT: Report type of gradient descent
    learning = batch_learning(num_of_files_in_path, batch_size)  #calculates which type of batch learning is being used
    print("%s files in path." %num_of_files_in_path ,"// Batch size =",batch_size, "\nLearning via: " + learning,"\n") #prints the number of files in the path and the batch size and the resultant type of batch learning

    #%% - Data Loader & Preperation

    ### - Dataset
    # Dataset loading function
    def train_loader2d(path): #loads the 2D image from the path
        sample = (np.load(path))
        if TORCHSIM_data:               ####REMOVE ONCE TSIM IS FIXED!!!
            sample = sample * 10    
        return (sample) 
    
    # Creates the training dataset using the DatasetFolder function from torchvision.datasets
    train_dataset = torchvision.datasets.DatasetFolder(train_dir, train_loader2d, extensions='.npy') 

    ### - Data Preparation 
    tensor_transform = partial(np_to_tensor, double_precision=double_precision) #using functools partial to bundle the args into np_to_tensor to use in custom torch transform using lambda function

    # Transformations
    train_transform = transforms.Compose([transforms.Lambda(tensor_transform),
                                        #transforms.RandomRotation(30),         #transforms.RandomRotation(angle (degrees?) ) rotates the tensor randomly up to max value of angle argument
                                        #transforms.RandomResizedCrop(224),     #transforms.RandomResizedCrop(pixels) crops the data to 'pixels' in height and width (#!!! and (maybe) chooses a random centre point????)
                                        #transforms.RandomHorizontalFlip(),     #transforms.RandomHorizontalFlip() flips the image data horizontally 
                                        ])      
    train_dataset.transform = train_transform   
    ##For info on all transforms check out: https://pytorch.org/vision/0.9/transforms.html

    # Dataset Partitioning
    m = len(train_dataset)  #m is the length of the train_dataset, i.e the number of images in the dataset
    train_split = int(m * train_test_split_ratio) #train_split is the ratio of train images to be used in the training set as opposed to non_training set
    train_data, non_training_data = torch.utils.data.random_split(train_dataset, [train_split, m-train_split])    #random_split(data_to_split, [size of output1, size of output2]) just splits the train_dataset into two parts, 4/5 goes to train_data and 1/5 goes to val_data , validation?

    m2 = len(non_training_data)  #m2 is the length of the non_training_data, i.e the number of images in the dataset
    if val_set_on:
        val_split = int(m2 * val_test_split_ratio) #val_split is the ratio of npon train images to be used in the validation set as opposed to test set
        test_data, val_data = torch.utils.data.random_split(non_training_data, [m2 - val_split, val_split])  
    else:   #this all needs cleaning up, dont need val set in this case bu tnot having one does break other lines
        test_data = non_training_data  
        val_data = non_training_data

    # Generating Dataloaders
    train_loader = torch.utils.data.DataLoader(train_data, batch_size=batch_size, shuffle=shuffle_train_data)   #Training data loader, is shuffled based on parameter 'shuffle_train_data' = True/False
    test_loader = torch.utils.data.DataLoader(test_data, batch_size=batch_size)                                 #Testing data loader 
    valid_loader = torch.utils.data.DataLoader(val_data, batch_size=batch_size)                                 #Validation data loader 


    #%% - Setup model, loss criteria and optimiser    
    # Initialize the encoder and decoder
    encoder = Encoder(latent_dim, print_encoder_debug, fc_input_dim)
    decoder = Decoder(latent_dim, print_decoder_debug, fc_input_dim)

    # Sets the encoder and decoder to double precision floating point arithmetic (fp64)
    if double_precision:
        encoder.double()   
        decoder.double()

    # Define the optimizer
    params_to_optimize = [{'params': encoder.parameters()} ,{'params': decoder.parameters()}] #Selects what to optimise, 
    optim = torch.optim.Adam(params_to_optimize, lr=learning_rate, weight_decay=optim_w_decay)

    # Load in pretrained weights and biases, if user requested
    if start_from_pretrained_model:

        # load the full state dictionary into memory
        full_state_dict = torch.load(pretrained_model_path)

        # load the state dictionaries into the models
        encoder.load_state_dict(full_state_dict['encoder_state_dict'])
        decoder.load_state_dict(full_state_dict['decoder_state_dict'])

        if load_pretrained_optimser:
            # load the optimizer state dictionary, if user requested
            optim.load_state_dict(full_state_dict['optimizer_state_dict'])

    # Initalise Model on compute device
    encoder.to(device)   #Moves encoder to selected device, CPU/GPU
    decoder.to(device)   #Moves decoder to selected device, CPU/GPU

    #%% - Prepare Network Summary
    # Set evaluation mode for encoder and decoder

    with torch.no_grad(): # No need to track the gradients
        
        # Create dummy input tensor
        enc_input_tensor = torch.randn(batch_size, channels, ydim, xdim) 
        if double_precision:
            enc_input_tensor = enc_input_tensor.double()
            
        # Join the encoder and decoder models
        full_network_model = torch.nn.Sequential(encoder, decoder)   #should this be done with no_grad?

        # Generate network summary and then convert to string
        model_stats = summary(full_network_model, input_data=enc_input_tensor, device=device, verbose=0)
        summary_str = str(model_stats)             

        # Print Encoder/Decoder Network Summary to terminal if user requested (Regardless of user setting, network summary is always added to the final .txt output file along with the final model)
        if print_network_summary:
            print(summary_str)


    #%% - Add Activation data hooks to encoder and decoder layers
    if record_activity or record_weights or record_biases:   # MOVE TO FUCNTION AND THEN TO HELPER FUCNS FILE FOR CLEANER CODE

        # Create a list to store the activations of each layer
        #all_activations = []
        #recorded_weights = [] 
        #recorded_biases = [] 

        # Loop through all the modules (layers) in the encoder and register the hooks
        for idx, module in enumerate(encoder.encoder_lin.modules()):
            if isinstance(module, torch.nn.Linear):
                print("Registering hooks for encoder layer: ", idx)
                if record_activity:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(activation_hook_fn, layer_index=idx))
                if record_weights:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(weights_hook_fn, layer_index=idx))
                if record_biases:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(biases_hook_fn, layer_index=idx))

        enc_max_idx = idx
        # Loop through all the modules (layers) in the decoder and register the hooks
        for idx, module in enumerate(decoder.decoder_lin.modules()):
            if isinstance(module, torch.nn.Linear):
                print("Registering hooks for decoder layer: ", enc_max_idx + idx)
                if record_activity:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(activation_hook_fn, layer_index = enc_max_idx + idx))
                if record_weights:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(weights_hook_fn, layer_index = enc_max_idx + idx))
                if record_biases:
                    # Register the hook with the layer_index as the key to identify activations for this layer
                    module.register_forward_hook(partial(biases_hook_fn, layer_index = enc_max_idx + idx))

        print("All hooks registered\n")




            





    #%% - Running Training Loop
    # this is a dictionary ledger of train val loss history
    #history_da={'train_loss':[], 'test_loss':[], 'val_loss':[]}   # Creates a variable called history_da which contains two lists, 'train_loss' and 'val_loss' which are both empty to start with. value are latter appeneded to the two lists by way of history_da['val_loss'].append(x)
    epoch_times_list = []                        # Creates a variable called epoch_times_list which is an empty list. values are latter appeneded to the list by way of epoch_times_list.append(x)

    # this is a dictionary ledger of train val loss history
    history_da['train_loss']  = []
    history_da['test_loss'] = []                   #Just creates a variable called history_da which contains two lists, 'train_loss' and 'val_loss' which are both empty to start with. value are latter appeneded to the two lists by way of history_da['val_loss'].append(x)


    print("\nTraining Initiated\nPress Ctr + c to exit and save the model during training.\n")
    start_time = time.time()                     # Begin the training timer

    try:    # Try except clause allows user to exit training gracefully whilst still retaiing a saved model and ouput plots

        if print_partial_training_losses:        # Prints partial train losses per batch
            loop_range = range(1, num_epochs+1)
        else:                                    # No print partial train losses per batch, instead create progress bar
            loop_range = tqdm(range(1, num_epochs+1), desc='Epochs', colour='red')
        
        epoch_start_time = time.time()           # Starts the epoch timer
        for epoch in loop_range:                 # For loop that iterates over the number of epochs where 'epoch' takes the values (0) to (num_epochs - 1)
            if print_partial_training_losses:
                print(f'\nStart of EPOCH {epoch + 1}/{num_epochs}')

            ###CLEAN UP THIS METHOD TO SOMTHING BETTER!!!!!!
            avg_loss_mse = []
            avg_loss_mae = []
            avg_loss_snr = []
            avg_loss_psnr = []
            avg_loss_ssim = []
            avg_loss_nmi = []
            avg_loss_cc = []
            avg_loss_true_positive_xy = []
            avg_loss_true_positive_tof = []
            avg_loss_false_positive_xy = []

            ### Training (use the training function)
            train_loss=train_epoch(encoder, 
                                    decoder, 
                                    device, 
                                    train_loader, 
                                    loss_fn, 
                                    optim,
                                    signal_points,
                                    noise_points,
                                    x_std_dev, 
                                    y_std_dev,
                                    tof_std_dev,
                                    time_dimension,
                                    reconstruction_threshold,
                                    print_partial_training_losses,
                                    masking_optimised_binary_norm,
                                    loss_vs_sparse_img)

            ### Testing (use the testing function)
            test_loss = test_epoch(encoder, 
                                    decoder, 
                                    device, 
                                    test_loader, 
                                    loss_fn,
                                    signal_points,
                                    noise_points,
                                    x_std_dev, 
                                    y_std_dev,
                                    tof_std_dev,
                                    time_dimension,
                                    reconstruction_threshold,
                                    print_partial_training_losses,
                                    masking_optimised_binary_norm,
                                    loss_vs_sparse_img)
            
            if print_partial_training_losses:
                print('\n End of EPOCH {}/{} \t train loss {:.3f} \t val loss {:.3f}'.format(epoch + 1, num_epochs, train_loss, val_loss))     #epoch +1 is to make up for the fact the range spans 0 to epoch-1 but we want to numerate things from 1 upwards for sanity
            
            if epoch % print_every_other == 0 and epoch != 0:                        
                # Run plotting function for training feedback and telemetry.
                encoder.eval()
                decoder.eval()

                returned_data_from_plotting_function = plot_epoch_data(encoder, 
                                                                        decoder, 
                                                                        epoch, 
                                                                        model_save_name, 
                                                                        time_dimension, 
                                                                        reconstruction_threshold,
                                                                        signal_points,
                                                                        noise_points,
                                                                        x_std_dev, 
                                                                        y_std_dev,
                                                                        tof_std_dev,
                                                                        physical_scale_parameters,
                                                                        num_to_plot)
                
                number_of_true_signal_points, number_of_recovered_signal_points, in_data, noisy_data, rec_data = returned_data_from_plotting_function
                
                if save_all_raw_plot_data:
                    save_variable(number_of_true_signal_points, f'Epoch {epoch}_number_of_true_signal_points', raw_plotdata_output_dir, force_pickle=True)
                    save_variable(number_of_recovered_signal_points, f'Epoch {epoch}_number_of_recovered_signal_points', raw_plotdata_output_dir, force_pickle=True)
                    save_variable(in_data, f'Epoch {epoch}_input_list', raw_plotdata_output_dir, force_pickle=True)
                    save_variable(noisy_data, f'Epoch {epoch}_noised_list', raw_plotdata_output_dir, force_pickle=True)
                    save_variable(rec_data, f'Epoch {epoch}_recovered_list', raw_plotdata_output_dir, force_pickle=True)
        
                encoder.train()
                decoder.train()
            
            epoch_end_time = time.time()
            epoch_time = epoch_end_time - epoch_start_time
            epoch_times_list.append(epoch_time) 
            #print("EPOCH TIM LIST SIZE", len(epoch_times_list))       

            # Update the epoch reached counter  
            max_epoch_reached = epoch    

            ###CLEAN UP THIS METHOD TO SOMTHING BETTER!!!!!!
            epoch_avg_loss_mse.append(np.mean(avg_loss_mse))
            epoch_avg_loss_mae.append(np.mean(avg_loss_mae))
            epoch_avg_loss_snr.append(np.mean(avg_loss_snr))
            epoch_avg_loss_psnr.append(np.mean(avg_loss_psnr))
            epoch_avg_loss_ssim.append(np.mean(avg_loss_ssim))
            epoch_avg_loss_nmi.append(np.mean(avg_loss_nmi))
            epoch_avg_loss_cc.append(np.mean(avg_loss_cc))
            epoch_avg_loss_true_positive_xy.append(np.mean(avg_loss_true_positive_xy))
            epoch_avg_loss_true_positive_tof.append(np.mean(avg_loss_true_positive_tof))
            epoch_avg_loss_false_positive_xy.append(np.mean(avg_loss_false_positive_xy))

            write_hook_data_to_disk_and_clear(activations, weights_data, biases_data, epoch, dir)

            history_da['train_loss'].append(train_loss)
            history_da['test_loss'].append(test_loss)
            #print("TRAIN LOSS SIZE", len(history_da['train_loss'])) 
            
            
            ### Live Comparison Plots
            if plot_comparative_loss:
                create_comparison_plot_data(slide_live_plot_size, epoch, max_epoch_reached, comparative_live_loss, comparative_loss_titles, comparative_epoch_times, comparative_history_da, data=history_da['train_loss'])

            ### Training Timeout Check
            if timeout_training:
                if time.time() - start_time > timeout_time * 60:  #convert timeout time from minuits to seconds
                    print("Training timed out, exiting training loop")
                    break

            pass # end of try clause, if all goes well and user doesen't request an early exit then the training loop will end here

    # If user presses Ctr + c to exit training loop, this handles the exception and allows the code to run its final data and model saving etc before exiting        
    except KeyboardInterrupt:
        print("Keyboard interrupt detected. Exiting training gracefully...")


    #%% - After Training
    encoder.eval()
    decoder.eval()    
    
    # Stop timing the training process and calculate the total training time
    end_time = time.time()
    training_time = end_time - start_time

    # Report the training time
    print(f"\nTotal Training Cycle Took {training_time:.2f} seconds")

    # Build Dictionary to collect all output data for .txt file
    full_data_output = {}
    full_data_output["Train Loss"] = round(history_da['train_loss'][-1], 3)
    full_data_output["Test Loss"] = round(history_da['test_loss'][-1], 3)   #Val loss calulaton is broken? check it above

    #%% - Save any remaining unsaved raw plot data
    epochs_range = range(1,len(history_da["train_loss"])+1) 
    if save_all_raw_plot_data:     ##FIND bETTER PLACE FOR THIS
        save_variable(np.array(epochs_range), "epochs_range", raw_plotdata_output_dir)  #!!!!!! testing raw plot outputter!!!!

        save_variable(history_da, "history_da", raw_plotdata_output_dir) 

        save_variable(epoch_times_list, "epoch_times_list", raw_plotdata_output_dir) 

        save_variable(settings, "settings", raw_plotdata_output_dir)

        detailed_performance_loss_dict={}
        detailed_performance_loss_dict["epoch_avg_loss_mse"] = epoch_avg_loss_mse
        detailed_performance_loss_dict["epoch_avg_loss_snr"] = epoch_avg_loss_snr
        detailed_performance_loss_dict["epoch_avg_loss_psnr"] = epoch_avg_loss_psnr
        detailed_performance_loss_dict["epoch_avg_loss_ssim"] = epoch_avg_loss_ssim
        detailed_performance_loss_dict["epoch_avg_loss_nmi"] = epoch_avg_loss_nmi
        detailed_performance_loss_dict["epoch_avg_loss_cc"] = epoch_avg_loss_cc
        detailed_performance_loss_dict["epoch_avg_loss_true_positive_xy"] = epoch_avg_loss_true_positive_xy
        detailed_performance_loss_dict["epoch_avg_loss_true_positive_tof"] = epoch_avg_loss_true_positive_tof
        detailed_performance_loss_dict["epoch_avg_loss_false_positive_xy"] = epoch_avg_loss_false_positive_xy
        save_variable(detailed_performance_loss_dict, "detailed_performance_loss", raw_plotdata_output_dir)


    #%% - Output Visulisations
    ###Loss function plots
    if plot_train_loss:
        Out_Label =  graphics_dir + f'{model_save_name} - Train loss - Epoch {epoch}.png'
        loss_plot(epochs_range, history_da['train_loss'], "Epoch number", f"Train loss ({loss_fn_label})", "Training loss", Out_Label, plot_or_save)

    if plot_test_loss:
        Out_Label =  graphics_dir + f'{model_save_name} - Val loss - Epoch {epoch}.png'
        loss_plot(epochs_range, history_da['test_loss'], "Epoch number", f"Test loss ({loss_fn_label})", "Test loss", Out_Label, plot_or_save)

    if plot_time_loss:
        Out_Label = graphics_dir + f'{model_save_name} - Train loss v Time - Epoch {epoch}.png'
        loss_plot(epoch_times_list, history_da['train_loss'], "Training Time (s)", f"Train loss ({loss_fn_label})", "Training Time v Loss", Out_Label, plot_or_save)

    if plot_detailed_performance_loss: 

        fig, axs = plt.subplots(nrows=3, ncols=3, figsize=(15, 15))
        #epochs_range= epochs_range[1:]
        axs[0, 0].plot(epochs_range, epoch_avg_loss_mse)
        axs[0, 0].set_title("MAE loss")
        axs[0, 0].set_xlabel("Epoch number")
        axs[0, 0].set_ylabel("Loss (MAE)")
        axs[0, 0].grid(alpha=0.2) 

        axs[0, 1].plot(epochs_range, epoch_avg_loss_snr)
        axs[0, 1].set_title("SNR loss")
        axs[0, 1].set_xlabel("Epoch number")
        axs[0, 1].set_ylabel("Loss (SNR)")
        axs[0, 1].grid(alpha=0.2) 

        axs[0, 2].plot(epochs_range, epoch_avg_loss_psnr)
        axs[0, 2].set_title("PSNR loss")
        axs[0, 2].set_xlabel("Epoch number")
        axs[0, 2].set_ylabel("Loss (PSNR)")
        axs[0, 2].grid(alpha=0.2) 

        axs[1, 0].plot(epochs_range, epoch_avg_loss_ssim)
        axs[1, 0].set_title("SSIM loss")
        axs[1, 0].set_xlabel("Epoch number")
        axs[1, 0].set_ylabel("Loss (SSIM)")
        axs[1, 0].grid(alpha=0.2) 

        axs[1, 1].plot(epochs_range, epoch_avg_loss_nmi)
        axs[1, 1].set_title("NMI loss")
        axs[1, 1].set_xlabel("Epoch number")
        axs[1, 1].set_ylabel("Loss (NMI)")
        axs[1, 1].grid(alpha=0.2) 

        axs[1, 2].plot(epochs_range, epoch_avg_loss_cc)
        axs[1, 2].set_title("Coreelation Coefficent? loss")
        axs[1, 2].set_xlabel("Epoch number")
        axs[1, 2].set_ylabel("Loss (CC)")
        axs[1, 2].grid(alpha=0.2) 

        axs[2, 0].plot(epochs_range, epoch_avg_loss_true_positive_xy)
        axs[2, 0].set_title("True Positive XY loss")
        axs[2, 0].set_xlabel("Epoch number")
        axs[2, 0].set_ylabel("Loss (True Positive XY %)")
        axs[2, 0].set_ylim(-5 ,105)
        axs[2, 0].grid(alpha=0.2) 

        axs[2, 1].plot(epochs_range, epoch_avg_loss_true_positive_tof)
        axs[2, 1].set_title("True Positive TOF loss")
        axs[2, 1].set_xlabel("Epoch number")
        axs[2, 1].set_ylabel("Loss (True Positive TOF %)")
        axs[2, 1].set_ylim(-5 ,105)
        axs[2, 1].grid(alpha=0.2) 

        axs[2, 2].plot(epochs_range, epoch_avg_loss_false_positive_xy)
        axs[2, 2].set_title("False Positive XY loss BROKEN?")
        axs[2, 2].set_xlabel("Epoch number")
        axs[2, 2].set_ylabel("Loss (False Positive XY)")
        axs[2, 2].grid(alpha=0.2)

        Out_Label = graphics_dir + f'{model_save_name} - Detailed Performance loss - Epoch {epoch}.png'
        plot_save_choice(plot_or_save, Out_Label)

    if plot_pixel_threshold_telemetry:
        plot_telemetry(telemetry, signal_points, plot_or_save=plot_or_save)

    if plot_pixel_difference:
        num_diff_noised, num_same_noised, num_diff_cleaned, num_same_cleaned, im_diff_noised, im_diff_cleaned = AE_visual_difference(in_data, noisy_data, rec_data)
        full_data_output["num_diff_noised"] = num_diff_noised
        full_data_output["num_same_noised"] = num_same_noised
        full_data_output["num_diff_cleaned"] = num_diff_cleaned
        full_data_output["num_same_cleaned"] = num_same_cleaned

        # Create a new figure
        fig, (ax1, ax2) = plt.subplots(nrows=1, ncols=2, figsize=(8, 4))

        # Plot the im_diff_noised subplot
        ax1.imshow(im_diff_noised, cmap='gray')
        ax1.set_title('Original -> Noised Img')

        # Plot the im_diff_cleaned subplot
        ax2.imshow(im_diff_cleaned, cmap='gray')
        ax2.set_title('Original -> Cleaned Img')

        # Add title for the whole figure
        fig.suptitle('Pixel-wise Difference Comparisons')
        
        Out_Label = graphics_dir + f'{model_save_name} - Pixel Difference Epoch {epoch}.png'
        plot_save_choice(plot_or_save, Out_Label)  

    if plot_latent_generations:
        def show_image(img):
            npimg = img.numpy()
            plt.imshow(np.transpose(npimg, (1, 2, 0)))

        img_recon = Generative_Latent_information_Visulisation(encoder, decoder, latent_dim, device, test_loader)
        
        fig, ax = plt.subplots(figsize=(20, 8.5))
        show_image(torchvision.utils.make_grid(img_recon[:100],10,10, pad_value=100))
        plt.axis("off")
        Out_Label = graphics_dir + f'{model_save_name} - Latent Generation Epoch {epoch}.png'
        plot_save_choice(plot_or_save, Out_Label)  

    if plot_higher_dim:
        encoded_samples, tsne_results = Reduced_Dimension_Data_Representations(encoder, device, train_dataset, plot_or_save=plot_or_save)
        
        # Higher dim
        plt.scatter(encoded_samples['Enc. Variable 0'], encoded_samples['Enc. Variable 1'],
                c=encoded_samples['label'], alpha=0.7)
        plt.grid()
        Out_Label = graphics_dir + f'{model_save_name} - Higher Dimensisions Epoch {epoch}.png'
        plot_save_choice(plot_or_save, Out_Label)  

        # TSNE of Higher dim
        plt.scatter(tsne_results[:, 0], tsne_results[:, 1], c=encoded_samples['label'])
        plt.xlabel('tsne-2d-one')
        plt.ylabel('tsne-2d-two')
        plt.grid()
        Out_Label = graphics_dir + f'{model_save_name} - TSNE Epoch {epoch}.png'
        plot_save_choice(plot_or_save, Out_Label)  

    if plot_normalised_radar:
        # Data for the radar plot
        categories = list(map(str, detailed_performance_loss_dict.keys())) # Names of each performance measure
        #data = HARD!!! #np.array(([3, 1, 74, 1, 1], [6, 52, 2, 76, 6], [-4, 272, 2, 1, 6]))  # List of models each with list of values for each measure
        multi_labels = [f"{model_save_name}"].extend(comparative_loss_titles)  # Labels for each plot
        #create_radar_plot_multi_model(categories, data, multi_labels)
           
    if plot_Graphwiz:
        Graphviz_visulisation(encoder, decoder, double_precision, batch_size, xdim, ydim, graphics_dir)
        #Out_Label = graphics_dir + f'{model_save_name} - Graphwiz Epoch {epoch}.png'    
        #plot_save_choice(plot_or_save, Out_Label)  
    #%% - Export models
    print("\nSaving model to .pth file in output dir...")
    # Save and export trained model to user output dir
    torch.save((encoder, decoder), full_model_path)
    print("- Completed -")

    print("\nSaving model state dictionary to output dir...")
    # Save the state dictionary
    torch.save({'encoder_state_dict': encoder.state_dict(),   
                'decoder_state_dict': decoder.state_dict(),   
                'optimizer_state_dict': optim.state_dict()},  
                full_statedict_path)  
    print("- Completed -") 



    print("\nSaving Autoencoder python file to output dir...")

    # Save the latent dimesion value for automatic setting when deploying the models
    save_variable(latent_dim, "deployment_variables_ld", model_output_dir)

    # save the fc_input_dim value for automatic setting when deploying the models
    save_variable(fc_input_dim, "deployment_variables_fc_input_dim", model_output_dir)

    # Save the processing precision value for automatic setting when deploying the models
    save_variable(str(double_precision), "deployment_variables_double_precision", model_output_dir)

    # Get the directory name of the current Python file for the autoencoder export
    search_dir = os.path.abspath(__file__)
    search_dir = os.path.dirname(search_dir)

    # Locate .py file that defines the Encoder and Decoder and copies it to the model save dir, due to torch.save model save issues
    AE_file_name = Robust_model_export(Encoder, search_dir, model_output_dir, debug=debug_model_exporter) #Only need to run on encoder as encoder and decoder are both in the same file so both get saved this way
    print("- Completed -")

    #%% - Export all data logged to disk in form of .txt file in the output dir
    print("\nSaving inputs, model stats and results to .txt file in output dir...")
    #Comparison of true signal points to recovered signal points
    try: 
        #print("True signal points",number_of_true_signal_points)
        #print("Recovered signal points: ",number_of_recovered_signal_points, "\n")
        full_data_output["true_signal_points"] = number_of_true_signal_points  # Save the number of true signal points to the full_data_output dictionary
        full_data_output["recovered_signal_points"] = number_of_recovered_signal_points  # Save the number of recovered signal points to the full_data_output dictionary
    except:
        pass
    
    # Save .txt Encoder/Decoder Network Summary
    with open(full_netsum_filepath, 'w', encoding='utf-8') as output_file:    #utf_8 encoding needed as default (cp1252) unable to write special charecters present in the summary
        # Write the local date and time to the file
        TD_now = datetime.datetime.now()         # Get the current local date and time
        output_file.write(f"Date data taken: {TD_now.strftime('%Y-%m-%d %H:%M:%S')}\n")     # Write the current local date and time to the file
        output_file.write(("Model ID: " + model_save_name + f"\nTrained on device: {device}\n"))   # Write the model ID and device used to train the model to the file
        output_file.write((f"Max Epoch Reached: {max_epoch_reached}\n"))  # Write the max epoch reached during training to the file
        output_file.write((f"Training Time: {format_time(training_time)}\n")) # Write the training time to the file

        output_file.write((f"\nTorch Seed Val to Reproduce Run: {T_seed}\n")) 
        output_file.write((f"Numpy Seed Val to Reproduce Run: {N_seed}\n")) 
        if seeding_value:
            output_file.write((f"User Set Seeding: True\n"))
        else:
            output_file.write((f"User Set Seeding: False\n"))
        
        output_file.write("\nInput Settings:\n")  # Write the input settings to the file
        for key, value in settings.items():
            output_file.write(f"{key}: {value}\n")
        
        output_file.write("\nLoss Function Settings:\n")    # Write the loss function settings to the file
        output_file.write(f"Loss Function Choice: {loss_fn}\n")
        if loss_function_selection == 0:
            output_file.write(f"zero_weighting: {zero_weighting}\n")
            output_file.write(f"nonzero_weighting: {nonzero_weighting}\n")    
        if loss_function_selection == 6:
            output_file.write(f"Split Loss Functions Selected (Hits, Non-Hits): {split_loss_functions}\n")

        output_file.write("\nNormalisation:\n")     # Write the normalisation settings to the file
        output_file.write(f"Mask optimised binary norm: {masking_optimised_binary_norm}\n") 

        output_file.write("\nPre Training:\n")   # Write the pre training settings to the file
        output_file.write(f"start_from_pretrained_model: {start_from_pretrained_model}\n")
        if start_from_pretrained_model:
            output_file.write(f"pretrained_model_path: {pretrained_model_path}\n")
            output_file.write(f"full_statedict_path: {full_statedict_path}\n")    
            output_file.write(f"load_pretrained_optimser: {load_pretrained_optimser}\n")  
        
        output_file.write("\n \nFull Data Readouts:\n") 
        for key, value in full_data_output.items(): 
            output_file.write(f"{key}: {value}\n") 

        output_file.write("\nPython Lib Versions:\n") # Write the python library versions to the file
        output_file.write((f"PyTorch: {torch.__version__}\n"))  
        output_file.write((f"Torchvision: {torchvision.__version__}\n"))     
        output_file.write((f"Numpy: {np.__version__}\n")) 
        output_file.write((f"Pandas: {pd.__version__}\n")) 
        #output_file.write((f"MPL: {plt.__version__}\n"))

        output_file.write("\nAutoencoder Network:\n")  # Write the autoencoder network settings to the file
        output_file.write((f"AE File ID: {AE_file_name}\n"))    # Write the autoencoder network file ID to the file
        output_file.write("\n" + summary_str)   # Write the autoencoder network summary to the file

        output_file.write("\nTraining System Information:\n") 
        system_information = get_system_information()  # Get the system information using helper function
        output_file.write("\n" + system_information)  # Write the system information to the file
    print("- Completed -")

    if optimise_hyperparameter:
        val_loss = validation_routine(  encoder, 
                                        decoder, 
                                        device, 
                                        valid_loader, 
                                        loss_fn,
                                        signal_points,
                                        noise_points,
                                        x_std_dev, 
                                        y_std_dev,
                                        tof_std_dev,
                                        time_dimension,
                                        reconstruction_threshold,
                                        print_partial_training_losses,
                                        masking_optimised_binary_norm,
                                        loss_vs_sparse_img)
                
        history_da['val_loss'].append(val_loss)
        history_da['HTO_val'].append(HTO_val)
        history_da['training_time'].append(training_time)

        if print_validation_results:
            print("\n# VALIDATION RESULTS #")
            print(f"{hyperparam_to_optimise} value: ", HTO_val)
            print("Train loss: ", train_loss)
            print("Test loss: ", test_loss)
            print("Val loss: ", val_loss)
            print(f"Total Training time: {format_time(training_time)}", )

        print("\n\n- Model Training Completed -\n \n \n \n") 

#%% - Plot the validation loss and training time for each hyperparameter value
if optimise_hyperparameter:
    history_da['Hyperparameter_to_optimise'] = hyperparam_to_optimise
    if plot_validation_loss:
        Out_Label =  results_output_path_1 + model_save_name + "_" + hyperparam_to_optimise + r" Optimisation\\" + f'{model_save_name} - Val loss - Epoch {epoch}.png'
        loss_plot(history_da['HTO_val'], history_da['val_loss'], hyperparam_to_optimise, "Val loss (MSE)", "Validation loss", Out_Label, plot_or_save)

    if plot_training_time:
        Out_Label =  results_output_path_1 + model_save_name + "_" + hyperparam_to_optimise + r" Optimisation\\" +  f'{model_save_name} - Training time - Epoch {epoch}.png'
        loss_plot(history_da['HTO_val'], history_da['training_time'], hyperparam_to_optimise, "Training time (s)", "Training time", Out_Label, plot_or_save)
    
    print("Hyperparameter optimisation complete.\nExtracting optimisation data to Excel file...")
    netsum_directory = results_output_path_1 + model_save_name + "_" + hyperparam_to_optimise + r" Optimisation\\"                 # set the directory path where the txt files are located
    #print("directory to grab txt data from, seems not to be working, check dir is correct !!!!!!!: ", netsum_directory)
    excel_output_path = netsum_directory + hyperparam_to_optimise + " Optimisation - Settings.xlsx"    # set the output path for the Excel file
    extract_data_to_excel(netsum_directory, excel_output_path)    
    colour_code_excel_file(excel_output_path)        
    print("Extraction complete\n \n")

    #%% Prepare analysis of output models perfromance

    # Run the full performance tests
    output_file_path = netsum_directory + 'Analysis\\'  # directory to save output to
    os.makedirs(output_file_path, exist_ok=True)
    model_names = []
    pretrained_model_folder_paths = []
    for i, HTOsetting in enumerate(history_da['HTO_val']):
        model_names.append(model_save_name + "_" + hyperparam_to_optimise + "_" + str(HTOsetting))
        pretrained_model_folder_paths.append(netsum_directory + str(HTOsetting) + f'\\{model_save_name} - Training Results\Model_Deployment\\')
    
    ### NOTE FIX!!!!!!!!!!!!!!!!!!!!!!!
    ##run_full_perf_tests(perf_analysis_num_files, perf_analysis_plot, True, perf_analysis_dataset_dir, output_file_path, model_names, pretrained_model_folder_paths, terminal_print=False)
    run_full_perf_tests(num_files=perf_analysis_num_files, 
                        plot=perf_analysis_plot, 
                        save_recovered_data=True, 
                        dataset_dir=perf_analysis_dataset_dir, 
                        output_file_path=output_file_path, 
                        model_names=model_names, 
                        pretrained_model_folder_paths=pretrained_model_folder_paths,  
                        debug_mode = debug_hpo_perf_analysis)




#%% - End of Program - Printing message to notify user!
print("\nProgram Complete - Shutting down...\n")    
    
