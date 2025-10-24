from tqdm import tqdm
import torch
import os
import torch
import matplotlib.pyplot as plt
import numpy as np
import random
import pickle
import pandas as pd
from Helper_files.DC3D_Core_Functions import masking_recovery, gaped_normalisation, gaped_renormalisation
#### NEED SFIXING
# ADD TIME DIM AS ARG
#  warn("File may not be readable: column headings must be strings.") 
# save is not part of the public API, usage can give unexpected results and will be removed in a future version       excel_writer.save()
# FIX THE BOXPLOTS X AXIS AND LEGEND NOT BEING FILLED WITH THE LABELS!?
# ADD TABLE STYLE TO TABLE!!! TABLE STYLE LIGHT 15!!
# Add the radar plots
 
time_dimension = 1000 ### FIX THIS SHOULD BE ARGUMENT FROM MAIN BODY!!!

#%% -  Helper Functions
def save_variable(variable, variable_name, path, force_pickle=False):

    if force_pickle:
        with open(path + variable_name + "_forcepkl.pkl", 'wb') as file:
            pickle.dump(variable, file)
    else:
        if isinstance(variable, dict):
            with open(path + variable_name + "_dict.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, np.ndarray):
            np.save(path + variable_name + "_array.npy", variable)

        elif isinstance(variable, torch.Tensor):
            torch.save(variable, path + variable_name + "_tensor.pt")

        elif isinstance(variable, list):
            df = pd.DataFrame(variable)
            df.to_csv(path + variable_name + "_list.csv", index=False)

        elif isinstance(variable, int):
            with open(path + variable_name + "_int.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, float):
            with open(path + variable_name + "_float.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, str):
            with open(path + variable_name + "_str.pkl", 'wb') as file:
                pickle.dump(variable, file)

        else:
            raise ValueError("Unsupported variable type.")

def plot_save_choice(plot_or_save, output_file_path):
    if plot_or_save == 0:
        plt.show()
    else:
        plt.savefig(output_file_path, format='png')    
        if plot_or_save == 1:    
            plt.close()
        else:
            plt.show()

#%% -  Performance Evaluation Functions
from skimage.metrics import peak_signal_noise_ratio, mean_squared_error, normalized_mutual_information, normalized_root_mse, structural_similarity, simple_metrics, variation_of_information, adapted_rand_error, tests
##### USE THE ABOUVE OVER THE BELOW AS THEY ARE MORE ACCURATE AND HAVE MORE OPTIONS FOR USE #####

#Signal to Noise Ratio (SNR)
def SNR(clean_input, noised_target):
    """
    Calculates the Signal to Noise Ratio (SNR) of a given signal and noise.
    SNR is defined as the ratio of the magnitude of the signal and the magnitude of the noise.
    
    Args:
    clean_input (torch.Tensor): The original signal.
    noised_target (torch.Tensor): The signal with added noise.
    
    Returns:
    The calculated SNR value.    
    """
    signal_power = torch.mean(torch.pow(clean_input, 2))

    noise = clean_input - noised_target 
    noise_power = torch.mean(torch.pow(noise, 2))

    snr = 10 * torch.log10(signal_power / noise_power)
       
    return (float(snr.numpy()))

#Peak Signal-to-Noise Ratio (PSNR):
def PSNR(clean_input, noised_target, time_dimension):
    """
    Calculates the Peak Signal to Noise Ratio (PSNR) of a given image and its recovered version. PSNR is defined as the ratio of 
    the maximum possible power of a signal and the power of corrupting noise. The measure focuses on how well high-intensity 
    regions of the image come through the noise, and pays much less attention to low intensity regions.

    Args:
    clean_input (torch.Tensor): The original image.
    noised_target (torch.Tensor): The recovered image.
    
    Returns:
    The calculated PSNR value.
    """
    mse = torch.mean(torch.pow(clean_input - noised_target, 2))   #Finds the mean square error
    max_value = time_dimension
    psnr = 10 * torch.log10((max_value**2) / mse)
    return (float(psnr.numpy()))

#Mean Squared Error (MSE):
def MSE(clean_input, noised_target):
    """
    Mean Squared Error (MSE)

    Args:
    clean_input (torch.Tensor): The original image.
    noised_target (torch.Tensor): The recovered image.
    
    Returns:
    The calculated Mean Squared Error value.
    """
    mse = torch.mean(torch.pow(clean_input - noised_target, 2))
    return (float(mse.numpy()))

#Mean Absolute Error (MAE):
def MAE(clean_input, noised_target):
    """
    Mean Absolute Error (MAE)

    Args:
    clean_input (torch.Tensor): The original image.
    noised_target (torch.Tensor): The recovered image.
    
    Returns:
    The calculated Mean Absolute Error value.
    """
    return float((torch.mean(torch.abs(clean_input - noised_target))).numpy())

#Structural Similarity Index (SSIM):
def SSIM(clean_input, noised_target):
    """
    Structural Similarity Index Measure (SSIM), is a perceptual quality index that measures the structural similarity between 
    two images. SSIM takes into account the structural information of an image, such as luminance, contrast, and structure, 
    and compares the two images based on these factors. SSIM is based on a three-part similarity metric that considers the 
    structural information in the image, the dynamic range of the image, and the luminance information of the image. SSIM is 
    designed to provide a more perceptually relevant measure of image similarity than traditional metrics such as Mean Squared 
    Error or Peak Signal-to-Noise Ratio.

    Args:
    clean_input (torch.Tensor): The original image.
    noised_target (torch.Tensor): The recovered image.
    
    Returns:
    The calculated Structural Similarity Index Measure value.
    """
    clean_image = clean_input.numpy()
    recovered_image = noised_target.numpy()
    return structural_similarity(clean_image, recovered_image, data_range=1.0)

#Correlation Coefficent
def correlation_coeff(clean_input, noised_target):
    
    """
    Correlation coefficient is a scalar value that measures the linear relationship between two signals. The correlation 
    coefficient ranges from -1 to 1, where a value of 1 indicates a perfect positive linear relationship, a value of -1 indicates 
    a perfect negative linear relationship, and a value of 0 indicates no linear relationship between the two signals. Correlation 
    coefficient only measures the linear relationship between two signals, and does not take into account the structure of the signals.

    ρ = cov(x,y) / (stddev(x) * stddev(y))

    The function first computes the mean and standard deviation of each tensor, and then subtracts the mean from each element 
    to get the centered tensors x_center and y_center. The numerator is the sum of the element-wise product of x_center 
    and y_center, and the denominator is the product of the standard deviations of the two centered tensors multiplied by the 
    number of elements in the tensor. The function returns the value of the correlation coefficient ρ as the ratio of the numerator 
    and denominator.

    Args:
    clean_input (torch.Tensor): The original image.
    noised_target (torch.Tensor): The recovered image.
    
    Returns:
    The calculated correlation coefficient value.
    """
    clean_mean = clean_input.mean()
    noised_mean = noised_target.mean()
    clean_std = clean_input.std()
    noised_std = noised_target.std()
    clean_center = clean_input - clean_mean
    noised_center = noised_target - noised_mean
    numerator = (clean_center * noised_center).sum()
    denominator = clean_std * noised_std * clean_input.numel()
    return float((numerator / denominator).numpy())

#Mutual Information:
def NomalisedMutualInformation(clean_input, noised_target):
    clean_image = clean_input.detach().cpu().numpy()
    recovered_image = noised_target.detach().cpu().numpy()
    return normalized_mutual_information(clean_image, recovered_image)-1

def compare_images_pixels(target_image, reconstructed_image, debug_mode=False): 
    """
    Takes in the clean image and the denoised image and compares them to find the percentages of signal spatial retention, signal temporal retention, and the raw count of false positives and false negatives.

    Args:
        target_image (torch tensor): The clean image
        reconstructed_image (torch tensor): The denoised image
        debug_mode (bool): If True, the function will print out the total number of pixels, the true signal points, and the true zero points in the target image. Default = False
        
    Returns:
        signal_spatial_retention_percentage (float): The percentage of signal spatial retention
        signal_temporal_retention_percentage (float): The percentage of signal temporal retention
        false_positive_count_raw (int): The raw count of false positives
        false_negative_count_raw (int): The raw count of false negatives
    """
    target_image = target_image.detach().cpu()
    reconstructed_image = reconstructed_image.detach().cpu()
    
    # determine the total number of pixels in the data
    total_pixels = target_image.numel()
    if debug_mode:
        print("Total Pixels: ", total_pixels)

    # Genrates a index mask for the zero values in the target image
    zero_mask = (target_image == 0)

    # Inverts the zero mask to get the non-zero mask from the target image indicating the signal points
    nonzero_mask = ~zero_mask

    # True Number of Signal Points in the Target Image
    true_signal_points = len(target_image[nonzero_mask])
    if debug_mode:
        print("True Signal Points: ", true_signal_points)

    # True Zero Points in the Target Image
    true_zero_points = len(target_image[zero_mask])
    if debug_mode:
        print("True Zero Points: ", true_zero_points)


    # detemine how many of the values in the reconstructed image that fall in the nonzero mask are non zero
    signal_spatial_retention_raw = len(reconstructed_image[nonzero_mask].nonzero())
    if debug_mode:
        print("Signal Spatial Retention Raw: ", signal_spatial_retention_raw)
    signal_spatial_retention_percentage = (signal_spatial_retention_raw / len(reconstructed_image[nonzero_mask])) * 100
    if debug_mode:
        print("Signal Spatial Retention Percentage: ", signal_spatial_retention_percentage, "%")

    # determine how many of those indexs have matching ToF values
    values_in_recon = reconstructed_image[nonzero_mask]
    values_in_target = target_image[nonzero_mask]
    signal_temporal_retention_raw = len(values_in_recon[values_in_recon == values_in_target])
    if debug_mode:
        print("Signal Temporal Retention Raw: ", signal_temporal_retention_raw)
    signal_temporal_retention_percentage = (signal_temporal_retention_raw / len(reconstructed_image[nonzero_mask])) * 100
    if debug_mode:
        print("Signal Temporal Retention Percentage: ", signal_temporal_retention_percentage, "%")


    # determine how many of the values in the reconstructed image that fall under the zero mask are not zero 
    false_positive_count_raw = len(reconstructed_image[zero_mask].nonzero())
    if debug_mode:
        print("False Positive Count: ", false_positive_count_raw)


    # Determine how many of the values in the reconstructed image that fall under the non zero mask are zero
    data = reconstructed_image[nonzero_mask]
    false_negative_count_raw = len(data[data == 0])
    if debug_mode:
        print("False negative count: ", false_negative_count_raw)

    return signal_spatial_retention_percentage, signal_temporal_retention_percentage, false_positive_count_raw, false_negative_count_raw, signal_spatial_retention_raw, signal_temporal_retention_raw


def image_loader(input_image_path):
    ### Load image from path 
    input_image = np.load(input_image_path)
    return input_image

#Combine all performance metrics into simple test script
def quantify_performance(clean_input, noised_target, label, debug_mode=False):

    if debug_mode:
        print("clean_input shape, type, dtype:", clean_input.shape, type(clean_input), clean_input.dtype)
        print("noised_target shape, type, dtype:", noised_target.shape, type(noised_target), noised_target.dtype)

    performance = {}
    performance['MSE'] = MSE(clean_input, noised_target)
    performance['MAE'] = MAE(clean_input, noised_target)
    performance['SNR'] = SNR(clean_input, noised_target)
    performance['PSNR'] = PSNR(clean_input, noised_target, time_dimension=1000)
    performance['SSIM'] = SSIM(clean_input, noised_target)
    performance['Normalised Mutual Information'] = NomalisedMutualInformation(clean_input, noised_target)    #BROKEN
    performance['Correlation Coefficient'] = correlation_coeff(clean_input, noised_target)
    percentage_of_true_positive_xy, percentage_of_true_positive_tof, numof_false_positives_xy, num_of_false_negative_x_y, signal_spatial_retention_raw, signal_temporal_retention_raw = compare_images_pixels(clean_input, noised_target)
    performance['Percentage of true pixels lit in recon'] = percentage_of_true_positive_xy
    performance['Percentage of true TOF pixels recon'] = percentage_of_true_positive_tof
    performance['Number of extra lit pixels in recon'] = numof_false_positives_xy
    #performance['Number of missing lit pixels in recon'] = num_of_false_negative_x_y   # ADD IN AND VERIFY
    # performance['Raw Signal Spatial Retention'] = signal_spatial_retention_raw
    # performance['Raw Signal Temporal Retention'] = signal_temporal_retention_raw
    return performance







 
#%% - Model Operation Functions
def initialise_model(pretrained_model_folder_path, verbose_mode=False): # Defines a function for initialising the model. Function takes inputs, 'encoder' and 'decoder' which are expected to be classes (defining the encode and decode nets), 'encoder_path' and 'decoder_path' which are the paths to the encoder and decoder weights, and 'device' which is the device the model is to be initialised on (CPU or GPU).
    import os

    if verbose_mode:
        print(f"Loading deployment files from: {pretrained_model_folder_path}")
    # Initialize variables to store the file paths
    py_file_path = None
    intpkl_file_path = None
    strpkl_file_path = None
    pth_model_file_path = None
    fc2pkl_file_path = None

    # Iterate through the files in the folder
    for file_name in os.listdir(pretrained_model_folder_path):
        file_path = os.path.join(pretrained_model_folder_path, file_name)
        
        if verbose_mode:
            print(f"Found file: {file_name}")
            
        if file_name.endswith('.py'):
            # Load the .py file
            py_file_path = file_path
            
        elif file_name.endswith('ld_int.pkl'):
            # Load the .pkl file
            intpkl_file_path = file_path

        elif file_name.endswith('dim_int.pkl'):
            # Load the .pkl file
            fc2pkl_file_path = file_path        

        elif file_name.endswith('_str.pkl'):
            strpkl_file_path = file_path
            
        elif file_name.endswith('State Dicts.pth'):
            # Load the .pth file with 'Model' in the filename
            pth_model_file_path = file_path

    # Check if all the required files are found
    if verbose_mode:
        if py_file_path and intpkl_file_path and strpkl_file_path and pth_model_file_path and fc2pkl_file_path:
            # Load the files into Python using their respective paths
            # Add your code here to process or use the files as needed
            print("Files loaded successfully!\n")
        else:
            print("One or more required files not found.\n")


        print(f"Loading settings from: {intpkl_file_path}, {fc2pkl_file_path} & {strpkl_file_path}")
    # Import the settings from the .pkl file
    import pickle
    with open(intpkl_file_path, 'rb') as f:
        latent_dim = pickle.load(f)
    latent_dim = int(latent_dim)

    if fc2pkl_file_path != None:
        with open(fc2pkl_file_path, 'rb') as f:
            fc_layer_size = pickle.load(f)
    
    else:
        fc_layer_size = 128
    fc_layer_size = int(fc_layer_size)

    with open(strpkl_file_path, 'rb') as f:
        double_precision = pickle.load(f)
    print("Settings loaded successfully!\n")

    # Import the encoder and decoder classes from the .py file
    def import_encoder_decoder(py_file_path):
        import importlib.util
        module_name = os.path.splitext(os.path.basename(py_file_path))[0]
        module_path = os.path.dirname(py_file_path)
        if verbose_mode:
            print("Module name: ", module_name + '.py')
            print("Module path: ", module_path)
        spec = importlib.util.spec_from_file_location(module_name + '.py', py_file_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module.Encoder, module.Decoder
    
    if verbose_mode:
        print(f"Loading Encoder and Decoder from: {py_file_path}")
    Encoder, Decoder = import_encoder_decoder(py_file_path)
 
    if fc2pkl_file_path != None:     
        encoder = Encoder(latent_dim, False, fc_layer_size)
        decoder = Decoder(latent_dim, False, fc_layer_size)
    else: #LEGACY !!! REMOVE ONCE ENOUGH NEW DATA
        encoder = Encoder(latent_dim, False)
        decoder = Decoder(latent_dim, False)

    # Sets the encoder and decoder to double precision floating point arithmetic (fp64)
    if double_precision:
        encoder.double()   
        decoder.double()

    
    if verbose_mode:
        print("Encoder and Decoder classes imported successfully!\n")

        print(f"Loading model weights from: {pth_model_file_path}")
    full_state_dict = torch.load(pth_model_file_path)                    # load saved model from disk 
    encoder.load_state_dict(full_state_dict['encoder_state_dict'])       # load the state dictionaries into the models
    decoder.load_state_dict(full_state_dict['decoder_state_dict'])
    if verbose_mode:
        print("Model weights loaded successfully!\n")

        print("Moving model to selected device:")
    device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
    if verbose_mode:
        print('Selected Tensor Processing Device: ' + f'{device}'.upper())  #Informs user if running on CPU or GPU

    encoder.eval()                                   #.eval() is a kind of switch for some specific layers/parts of the model that behave differently during training and inference (evaluating) time. For example, Dropouts Layers, BatchNorm Layers etc. You need to turn off them during model evaluation, and .eval() will do it for you. In addition, the common practice for evaluating/validation is using torch.no_grad() in pair with model.eval() to turn off gradients computation
    decoder.eval()    

    encoder.to(device)   #Moves encoder to selected device, CPU/GPU
    decoder.to(device)   #Moves decoder to selected device, CPU/GPU
    if verbose_mode:
        print("Model moved to device successfully!\n")

    return encoder, decoder, double_precision

def run_model(input_image_tensor, reconstruction_threshold, encoder, decoder, time_dimension=100):                                                        # Defines a function for generating a batch of test outputs from the autoencoder. And also the input + clean training data? Function takes inputs, 'encoder' and 'decoder' which are expected to be classes (defining the encode and decode nets), 'n' which is the number of ?????Images in the batch????, and 'noise_factor' which is a multiplier for the magnitude of the added noise allowing it to be scaled in intensity.  
    with torch.no_grad():
        norm_image = gaped_normalisation(input_image_tensor.clone(), reconstruction_threshold, time_dimension)
        image_prepared = norm_image.unsqueeze(0).unsqueeze(0)   #Adds two extra dimesnions to start of array so shape goes from (x,y) to (1,1,x,y) to represent batch and channel dims
        rec_image = decoder(encoder(image_prepared))                         #Creates a recovered image (denoised image), by running a noisy image through the encoder and then the output of that through the decoder.
        rec = rec_image.squeeze().numpy()
        rec_image_renorm = gaped_renormalisation(rec, reconstruction_threshold, time_dimension)
        masking_rec_image = masking_recovery(input_image_tensor.numpy(), rec_image_renorm, time_dimension, print_result=False)

    return rec_image_renorm, masking_rec_image

#%% -  Testing Functions
def test_single_model(number_of_files, pretrained_model_folder_path, dataset_dir, output_file_path, plot, save_recovered_data, model_name, debug_mode, file_list):
    encoder, decoder, double_precision = initialise_model(pretrained_model_folder_path, verbose_mode=debug_mode)
    if debug_mode:
        print("ENCODER & DECODER LOADED")

    # Create dictionary 
    files_dict_direct = {}
    files_dict_masked = {}

    ### Loop through the files and run function on each and collect image recovery stats
    for i, file_name in tqdm(enumerate(file_list), desc='Processing Files', leave=False, colour='purple'):
        image = image_loader(dataset_dir + '\\Labels\\' + file_name)  
        sparse_img = image_loader(dataset_dir + '\\Labels_Sparse\\' + file_name)
        noisy_image = image_loader(dataset_dir + '\\Data\\' + file_name)
        noisy_image_tensor = torch.tensor(noisy_image)
        noisy_image_batched = noisy_image_tensor      

        recovered, image_masked_rec = run_model(noisy_image_batched, reconstruction_threshold=0.5, encoder=encoder, decoder=decoder, time_dimension=1000)

        if save_recovered_data:
            np.save(output_file_path + f'Recovered_data\Direct\{model_name} {file_name}', np.array(recovered))
            np.save(output_file_path + f'Recovered_data\Masking\{model_name} {file_name}', np.array(image_masked_rec))
        
        #run stats collection functions
        if debug_mode:
            print("STATS COLLECTION STARTED")
        direct = quantify_performance(torch.tensor(image), torch.tensor(recovered), "Direct Recovery", debug_mode)
        
        # NOTE: CARE ! MASKING IS VALIDATED VS SPARSE IMG AND DIRECT IS VS FULL PATH LABEL, THEY ARE NOT ATALL COMPARABLE BUT NEEDS RESOLOUTION
        masking = quantify_performance(torch.tensor(sparse_img), torch.tensor(image_masked_rec), "Masking", debug_mode)
        if debug_mode:
            print("COMPLETE")

            print("PLOTTING STARTED")
        if plot != False:
            if i < plot:
                fig, axes = plt.subplots(1, 4, figsize=(14,4))
                axes[0].imshow(image)
                axes[0].set_title('Input Image')

                axes[1].imshow(noisy_image)
                axes[1].set_title('Noisy Image')

                axes[2].imshow(recovered)
                axes[2].set_title('Recovered Image')

                axes[3].imshow(image_masked_rec)
                axes[3].set_title('Masked Rec')

                Out_Label = output_file_path + f'Comparison_Images\{model_name} {file_name[:-4]}.png'
                plot_save_choice(1, Out_Label)   
        if debug_mode:
            print("COMPLETE")
            
        # add the results to the dictionary with the model name as the key
        files_dict_direct[file_name] = direct
        files_dict_masked[file_name] = masking

    return (files_dict_direct, files_dict_masked)

#%% - Plotting and Analysis Functions
# Turns results dictionary into a fully formatted excel file
def create_excel_file_and_calculate_stats(data_dictionary, output_file_path, debug=False):
    ### Excel file creator for validation results ###
    import pandas as pd

    # Create an Excel writer object
    excel_writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter')

    # Iterate through the nested dictionary and create sheets, rows, and columns
    for sheet_name, sheet_data in data_dictionary.items():
        df = pd.DataFrame.from_dict(sheet_data, orient='index')
        df.to_excel(excel_writer, sheet_name=sheet_name)

    # Save the Excel file
    excel_writer.save()    #####ERROR !!! .save is not in public API ad in future will be removed! FIX FIX FIX

    if debug:
        print(f"Excel file  has been created at: '{output_file_path}'")


    ### Excel file formatter and statistics for validation results ###
    # NOTE: Works great but excel will give error on opening as it has to fix some problem to do with the table formatting. i think its beacuse px only supports manually providing coloum headers unlike from inside excel natively where you can specify the table already has headers. once yuo accept the excel error message the file works perfectly including the tables. i assume it is finxing whatever issue there is automatically.

    input_file_path = output_file_path

    import openpyxl as px
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    # Load the Excel file
    workbook = px.load_workbook(input_file_path)

    ### File by File Results Sheets ###
    # Loop through each sheet in the Excel file
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        row_position_for_avg_cell = ws.max_row + 1
        num_of_files = ws.max_row - 1
        num_of_columns = ws.max_column

        # Iterate through each column and calculate the average ad statistics and add them to the sheet
        for col_num in range(1, ws.max_column + 1):

            column = ws[chr(64 + col_num)]
            
            # Calculate the cell for the AVERAGE function
            average_cell = f"{chr(64 + col_num)}{row_position_for_avg_cell}"
            std_dev_cell = f"{chr(64 + col_num)}{row_position_for_avg_cell + 1}"
            std_err_cell = f"{chr(64 + col_num)}{row_position_for_avg_cell + 2}"
            conf_int_cell = f"{chr(64 + col_num)}{row_position_for_avg_cell + 3}"

            if col_num == 1:
                #Insert text into cell indicating average
                ws[average_cell] = "MEAN"
                ws[std_dev_cell] = "STDEV"
                ws[std_err_cell] = "SEM"
                ws[conf_int_cell] = "CI"

            else:
                # Insert the AVERAGE function into the cell
                ws[average_cell] = f"=AVERAGE({chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1})"
                ws[std_dev_cell] = f"=STDEV({chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1})"
                ws[std_err_cell] = f"=STDEV({chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1})/SQRT(COUNT({chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1}))"
                ws[conf_int_cell] = f"=CONFIDENCE(0.05, STDEV({chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1}), {num_of_files})"

            # Add the conditional formatting to the column data
            # check if coloumn is even or odd
            if col_num % 2 == 0:
                column_colour = px.styles.colors.Color(rgb="CCFFFF")
            else:
                column_colour = px.styles.colors.Color(rgb="E5CCFF")

            # Create the data bars rule with solid fill
            data_bars_rule = DataBarRule(
                start_type='min',
                end_type='max',
                color=column_colour,  # Solid fill color
            )
        
            # Add the data bars rule to the column
            ws.conditional_formatting.add(f"{chr(64 + col_num)}2:{chr(64 + col_num)}{row_position_for_avg_cell-1}", data_bars_rule)


        #######ADD TABLE STYLE TO TABLE!!! TABLE STYLE LIGHT 15!!

        # Add the table formatting to the sheet
        #table_range = f"A{1}:{chr(64 + num_of_columns)}{1 + num_of_files}"    # Remove spaces from sheet name to create a valid table name
        table_range = f"A{1}:{chr(64 + num_of_columns)}{1 + num_of_files}"

        table_name = f"Table_{sheet.replace(' ', '_')}"
        table = px.worksheet.table.Table(displayName=table_name, ref=table_range)

        # Specify that the table has headers
        table.tableStyleInfo = px.worksheet.table.TableStyleInfo(showFirstColumn=True)

        # Add the table to the worksheet
        ws.add_table(table)
        


    ### Model by Model Results Sheet ###
    # Add a new sheet to the Excel file called Model by Model Results
    model_by_model_ws = workbook.create_sheet("Model by Model Results")

    # Duplicate the column headers from the first sheet to thsi sheet
    for col_num in range(1, ws.max_column + 1):
        model_by_model_ws[f"{chr(64 + col_num)}1"] = ws[f"{chr(64 + col_num)}1"].value

    # Target sheet for copy
    target_sheet_name = "Model by Model Results"
    model_by_model_ws = workbook[target_sheet_name]

    # Loop through the source sheets
    for sheet_name in workbook.sheetnames:
        if sheet_name != target_sheet_name:
            source_sheet = workbook[sheet_name]
            target_row = model_by_model_ws.max_row + 1  # Find the next empty row in the target sheet
            
            # Define the row number from which to copy values
            source_row_number = row_position_for_avg_cell  
            
            # Write the name of the source sheet into the first cell of the target row
            model_by_model_ws[f"A{target_row}"] = sheet_name
            
            for col_num in range(2, source_sheet.max_column + 1):
                source_cell = source_sheet.cell(row=source_row_number, column=col_num)

                # Extracting the cell range
                cell_range = (source_cell.value).split('(')[1].split(')')[0]

                # Reparsing the equation to refrence the original sheet name
                result = f"=AVERAGE('{source_sheet.title}'!{cell_range})"
                target_cell = model_by_model_ws.cell(row=target_row, column=col_num, value=result)
            
    # in sheet model by model results, add conditional formatting to the columns
    for col_num in range(2, model_by_model_ws.max_column + 1):
        column = model_by_model_ws[chr(64 + col_num)]
        # check if coloumn is even or odd
        if col_num % 2 == 0:
            column_colour = px.styles.colors.Color(rgb="CCFFFF")
        else:
            column_colour = px.styles.colors.Color(rgb="E5CCFF")

        # Create the data bars rule with solid fill
        data_bars_rule = DataBarRule(
            start_type='min',
            end_type='max',
            color=column_colour,  # Solid fill color
        )

        # Add the data bars rule to the column
        model_by_model_ws.conditional_formatting.add(f"{chr(64 + col_num)}2:{chr(64 + col_num)}{model_by_model_ws.max_row}", data_bars_rule)


        #######ADD TABLE STYLE TO TABLE!!! TABLE STYLE LIGHT 15!!
    # Add the table formatting to the sheet model by model sheet 
    table_range = f"A{1}:{chr(64 + num_of_columns)}{model_by_model_ws.max_row}"
    table_name = f"Table_Model_by_Model"
    table = px.worksheet.table.Table(displayName=table_name, ref=table_range)


    # Specify that the table has headers
    table.tableStyleInfo = px.worksheet.table.TableStyleInfo(showFirstColumn=False)

    # Add the table to the worksheet
    model_by_model_ws.add_table(table)


    # Save the modified Excel file
    workbook.save(output_file_path)

    if debug:
        print(f"Modified Excel file saved as {output_file_path}")

# Turns results dictionary into set of detailed plots 
def plot_multimodel_results(multi_model_results_dict, save_recovered_data, output_file_path, debug_mode=False):
    #print(multi_model_results_dict.keys())

    # get to third level in dictionary to get metric names, (this needs a better apprach this was fastest way i coudl think for now, dlevels 1 and 2 dont matter as level 3 keys are ssaem for all level 1 and 2 keys)
    random_key1 = random.choice(list(multi_model_results_dict.keys()))
    #print("random_key1:", random_key1)
    random_key2 = random.choice(list(multi_model_results_dict[random_key1].keys()))
    #print("random_key2:", random_key2)
    
    # Extract all the third-level keys 
    metrics = list(multi_model_results_dict[random_key1][random_key2].keys())
    #print("metrics:", metrics)


    colors = plt.cm.viridis(np.linspace(0, 1, len(list(multi_model_results_dict.keys()))))
    #print("colors:", colors )
    for metric in metrics:
        # Define a list of model names
        model_names = list(multi_model_results_dict.keys())
        #print("model_names:", model_names)
        # Define a list of file names
        file_names = [file_name for _, stats_dict in multi_model_results_dict.items() for file_name, perf_dict in stats_dict.items()]
        #print("file_names:", file_names)
        # Define a list of metric values for each model and file
        metric_values = [[perf_dict[metric] for file_name, perf_dict in stats_dict.items()] for _, stats_dict in multi_model_results_dict.items()]
        #print("metric_values:", metric_values)  
        # Create a box plot
        plt.boxplot(metric_values)
        
        # remove x ticks
        plt.xticks([])

        # Add scatter plot
        for i, coords in enumerate(metric_values):
            #print(model_names[i])
            x_coords = [i+1] * len(coords)
            plt.scatter(x_coords, coords, color=colors[i], alpha=0.7, label=model_names[i])
        
        # Set plot title
        plt.title(metric)

        # Set the y-axis label
        plt.ylabel(metric)

        # Show legend
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)

        # Show the plot
        plt.grid(alpha=0.2)

        Out_Label = output_file_path + f'Images\{metric}.png'
        plot_save_choice(1, Out_Label)           
        if debug_mode:
            print(f"Plot saved to '{Out_Label}'.")


#%% - Main code
def run_full_perf_tests(num_files, plot, save_recovered_data, dataset_dir, output_file_path, model_names, pretrained_model_folder_paths, debug_mode = False):
    # Create output path if not already in existance
    os.makedirs(output_file_path + 'Recovered_data\\Direct\\', exist_ok=True)
    os.makedirs(output_file_path + 'Recovered_data\\Masking\\', exist_ok=True)
    os.makedirs(output_file_path + 'Images\\', exist_ok=True)
    os.makedirs(output_file_path + 'Comparison_Images\\', exist_ok=True)
    os.makedirs(output_file_path + 'Raw_performance_data\\', exist_ok=True)

    # Create a dictionary to store the results for each model
    multi_model_results_dict = {} # create an empty dictionary to store the results for each model

    # Create file list of all the file names from the \Data folder
    file_list = [f for f in os.listdir(dataset_dir + '\\Data') if f.endswith('.npy')][:num_files]

    # Loop through each model and run the test
    for i, model_name in tqdm(enumerate(model_names), desc="Testing Model", colour='pink', leave=False):
        print(f"Running Model: {model_name}")
        files_dict_direct, files_dict_masked = test_single_model(num_files, pretrained_model_folder_paths[i], dataset_dir, output_file_path, plot, save_recovered_data, model_name, debug_mode, file_list)
        multi_model_results_dict[model_name + ' direct'] = files_dict_direct
        multi_model_results_dict[model_name + ' masked'] = files_dict_masked
    print("Data Gathering Completed")

    print("Analysing Results")
    # Save the results
    if save_recovered_data:
        save_variable(multi_model_results_dict, 'multi_model_results_dict', output_file_path + 'Raw_performance_data\\')
        if debug_mode:
            print("Raw Data Dictionary Saved")

    # Export the results to custom formatted excel file
    excel_file_path = output_file_path + 'Raw_performance_data\\' + 'Hyperparameter Validation Results.xlsx'
    create_excel_file_and_calculate_stats(multi_model_results_dict, excel_file_path, debug=False)    
    
    # Plots the results
    plot_multimodel_results(multi_model_results_dict, save_recovered_data, output_file_path)
    print("Analysis Finished\n\n")


#%% - Driver
"""


run_full_perf_tests(num_files=10, 
                     plot=10, 
                     save_recovered_data=True, 
                     dataset_dir=(r"N:\\Yr 3 Project Datasets\\PERF VALIDATION SETS\\10K 100N 30S\\"), 
                     output_file_path=r'Remove\\', 
                     model_names=["RDT 500K 1000ToF timed", "RDT 500K 1000ToF timed2"], 
                     pretrained_model_folder_paths=[r'N:\Yr 3 Project Results\RDT 500K 1000ToF timed - Training Results\Model_Deployment\\', r'N:\Yr 3 Project Results\RDT 500K 1000ToF timed - Training Results\Model_Deployment\\'],  
                     debug_mode = False)
"""