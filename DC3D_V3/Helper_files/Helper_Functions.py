import torch
import numpy as np
import pickle
import os
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

def set_model_precision(encoder, decoder, precision):
    if precision == 16: # Set encoder and decoder to half precision floating point arithmetic (fp16)
        encoder.half()
        decoder.half()
        dtype = torch.float16

    elif precision == 32: # Sets the encoder and decoder to single precision floating point arithmetic (fp32)
        encoder.float()
        decoder.float()
        dtype = torch.float32

    # Sets the encoder and decoder to double precision floating point arithmetic (fp64)
    elif precision == 64:
        encoder.double()   
        decoder.double()
        dtype = torch.float64

    else:
        raise ValueError("Error: Precision not set correctly, please set precision to 16, 32, or 64")

    return encoder, decoder, dtype

import os
import re

def extract_model_id_from_path(model_path):
    """
    Extracts the model ID from the given model path, regardless of where it appears.
    
    The model ID pattern is always 'MID<NUMBERS>__' and can be part of 
    any folder name or the filename.
    
    Args:
        model_path (str): The full file path of the model.
    
    Returns:
        int: The extracted model ID. Returns 0 if the model predates the model ID system.
    """
    # Search for a pattern like 'MID4__' or 'MID123__' anywhere in the path
    match = re.search(r'MID(\d+)__', model_path)
    
    if match:
        return int(match.group(1))
    else:
        # Model predates the model ID system
        ##! Could log or warn here
        return 0

def format_time(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours)}h :{int(minutes)}m :{seconds:3f}s"

# Helper function to return the batch learning method string to user
def batch_learning(training_dataset_size, batch_size):
    if batch_size == 1: 
        output = "Stochastic Gradient Descent"
    elif batch_size == training_dataset_size:
        output = "Batch Gradient Descent"        
    else:
        output = "Mini-Batch Gradient Descent"
    return(output)

def load_comparative_data(comparative_loss_paths, plot_live_training_loss=False, plot_live_time_loss=False):
    comparative_history_da = []
    comparative_epoch_times = []

    for loss_path in comparative_loss_paths:
        # load pkl file into dictionary
        if plot_live_training_loss or plot_live_time_loss:
            with open(loss_path + '\\Raw_Data_Output\\history_da_dict.pkl', 'rb') as f:
                comparative_history_da.append(pickle.load(f))
        
        if plot_live_time_loss:
            with open(loss_path + '\\Raw_Data_Output\\epoch_times_list_list.csv', 'rb') as f:
                # load the data from the csv file called f into a list 
                comparative_epoch_times.append(np.loadtxt(f, delimiter=',').tolist())
    
    return comparative_history_da, comparative_epoch_times

def save_variable(variable, variable_name, path, force_pickle=False):

    if force_pickle:
        with open(path + variable_name + "_forcepkl.pkl", 'wb') as file:
            pickle.dump(variable, file)
    else:
        if isinstance(variable, dict):
            with open(path + variable_name + "_dict.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, np.ndarray):
            np.save(path + variable_name + "_array.npy", variable)

        elif isinstance(variable, torch.Tensor):
            torch.save(variable, path + variable_name + "_tensor.pt")

        elif isinstance(variable, list):
            df = pd.DataFrame(variable)
            df.to_csv(path + variable_name + "_list.csv", index=False)

        elif isinstance(variable, int):
            with open(path + variable_name + "_int.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, float):
            with open(path + variable_name + "_float.pkl", 'wb') as file:
                pickle.dump(variable, file)

        elif isinstance(variable, str):
            with open(path + variable_name + "_str.pkl", 'wb') as file:
                pickle.dump(variable, file)

        else:
            raise ValueError("Unsupported variable type.")


# Helper function to allow values to be input as a range from which random values are chosen each time unless the input is a single value in which case it is used as the constant value
def input_range_to_random_value(*args):
    """
    Generate random values based on input ranges or single values.

    This function accepts an arbitrary number of arguments, where each argument
    can be either a single value (int or float) or a range (list or tuple) of
    values. For ranges, it generates a random integer if the range consists of
    integers, or a random float if the range consists of floats.

    Parameters:
    *args : int, float, list, tuple
        Arbitrary number of input arguments. Each argument can be a single value
        or a range represented as a list or tuple of two values.

    Returns:
    list
        A list containing the generated random values or the input values if they
        are already single values. If an input argument is not recognized as a
        value or range, None is appended to the list and an error message is printed.
    """
    results = []

    for input_range in args:

        if isinstance(input_range, (int, float)):
            ## If input is single value it is not randomised as is manually set
            results.append(input_range)
        
        elif isinstance(input_range, (list, tuple)):
            ## If input is a list or tuple then it is considered a range of values and is randomised in that range
            
            if all(isinstance(x, int) for x in input_range):
                ## If all values in the list are ints then the whole list is considered to be a range of ints and an int is returned
                results.append(torch.randint(input_range[0], input_range[1] + 1, (1,)).item())

            elif (isinstance(x, float) for x in input_range):
                ## Else if any single value in the list is a float then function will return a float
                results.append(torch.rand(1).item() * (input_range[1] - input_range[0]) + input_range[0])
            
            else:
                print("Error: input_range_to_random_value() input is not a value or pair of values in recognised format, must be float or int")
                results.append(None)
        else:
            print("Error: input_range_to_random_value() input is not a value or pair of values")
            results.append(None)
    
    return results

def colour_code_excel_file(file_path):
    """

    This function will take an excel file and colour code the cells based on the value of the cell. The colour coding is based on a colour scale from blue to red with blue being the lowest value and red being the highest value. The colour scale is applied to each column individually.

    Args:
        file_path (str): The path to the excel file to be colour coded
    
    Generates:
        Color coding in the excel file which is then saved back to disk in place of the original file
    """


    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Define the color scale rule for the conditional formatting
    color_scale = DataBarRule(
        start_type='num',
        start_value=0,
        end_type='num',
        end_value=1,
        color='0000FF',
        showValue=True,
    )

    # Apply the color scale rule to each column
    for column in sheet.columns:
        column_letter = get_column_letter(column[0].column)
        column_range = f'{column_letter}2:{column_letter}{sheet.max_row}'  # Assuming data starts from row 2
        column_cells = list(column)[1:]  # Skip the header cell

        for cell in column_cells:
            cell.number_format = '0.00'  # Optional: Format the cell as desired

        sheet.conditional_formatting.add(column_range, color_scale)

    #resave the workbook to the same file
    workbook.save(file_path)


# Helper function to clean up repeated plot save/show code
def plot_save_choice(plot_or_save, output_file_path=None, dpi=None):
    """
    Function used to set the save/display behavior for all figures and graphs generated by the program.

    Args:
        plot_or_save (int): A flag to set if the plot is printed to terminal or saved to disk. 
            0 prints plots to the terminal (blocking till closed).
            1 saves all end-of-epoch printouts to disk (non-blocking).
            2 saves outputs while also printing for the user (blocking till closed).
            3 neither saves nor shows any plots, immediately closed (useful for debugging).
        output_file_path (str): The path to save the plot to. Only needed if plot_or_save is set to 1 or 2.
    """
    if plot_or_save == 0 or plot_or_save == "show":
        plt.show()
    elif plot_or_save == 1 or plot_or_save == "save":  
        if dpi is not None:
            plt.savefig(output_file_path, format='png', dpi=dpi)
        else:
            plt.savefig(output_file_path, format='png')     
        plt.close()
    
    elif plot_or_save == 2 or plot_or_save == "both":
        if dpi is not None:
            plt.savefig(output_file_path, format='png', dpi=dpi)
        else:
            plt.savefig(output_file_path, format='png')   
        plt.show()
    elif plot_or_save == 3 or plot_or_save == "none":
        plt.close()
    else:
        raise ValueError("Invalid 'plot_or_save' value set for 'plot_save_choice' function call. "
                         "Please set to 0, 1, 2, or 3.\n"
                         "0 prints plots to terminal (blocking till closed).\n"
                         "1 saves all plots to disk (non-blocking).\n"
                         "2 prints to terminal and saves to disk (blocking till closed).\n"
                         "3 will neither save nor show any plots, they will be immediately closed, useful for debugging.")
    


# Getting model parameters for logging

def get_learning_rate(optimizer):
    for param_group in optimizer.param_groups:
        return param_group['lr']
    
def get_optimiser_parameters(optimizer):
    optim_params = {}
    for param_group in optimizer.param_groups:
        for key, value in param_group.items():
            if key != 'params':  # Skip the 'params' key as it contains the model parameters
                optim_params[key] = value
    return optim_params


    
def get_gradient_norm(encoder, decoder):
    total_norm = 0.0
    for p in list(encoder.parameters()) + list(decoder.parameters()):
        if p.grad is not None:
            param_norm = p.grad.data.norm(2)
            total_norm += param_norm.item() ** 2
    total_norm = total_norm ** (1. / 2)
    return total_norm   



def get_grad_and_param_norms(*models, norm_type: float = 2.0, eps: float = 1e-12):
    """
    Function to calulate the gradient and parameter norms for a list of models.

    Args:
        *models: A list of PyTorch models (e.g., encoder, decoder).
        norm_type (float): The type of norm to compute (default is 2 for L2 norm).
        eps (float): A small value to prevent division by zero (default is 1e-12).  
    
    Returns (grad_norm, param_norm) aggregated (default L2) across provided models.


    """
    total_grad = 0.0
    total_param = 0.0
    for m in models:
        for p in m.parameters():
            if p.grad is not None:
                total_grad += (p.grad.detach().norm(norm_type).item() ** norm_type)
            total_param += (p.detach().norm(norm_type).item() ** norm_type)
    grad_norm = (total_grad ** (1.0 / norm_type)) if total_grad > 0 else 0.0
    param_norm = (total_param ** (1.0 / norm_type)) if total_param > 0 else eps
    return float(grad_norm), float(param_norm)